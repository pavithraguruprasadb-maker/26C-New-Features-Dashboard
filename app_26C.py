import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import io

st.set_page_config(
    page_title="26C New Features Report Agent",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
    .main { background-color: #f8f9fa; }
    .stApp { font-family: 'Segoe UI', sans-serif; }
    .report-title { font-size: 2rem; font-weight: 700; color: #1a1a2e; margin-bottom: 0.2rem; }
    .report-subtitle { font-size: 1rem; color: #6c757d; margin-bottom: 1.5rem; }
    .metric-card {
        background: white; border-radius: 12px; padding: 1.2rem; text-align: center;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08); border-left: 4px solid #4f46e5;
    }
    .metric-value { font-size: 2rem; font-weight: 700; color: #4f46e5; }
    .metric-label { font-size: 0.85rem; color: #6c757d; margin-top: 4px; }
    .section-header {
        font-size: 1.1rem; font-weight: 600; color: #1a1a2e;
        margin: 1rem 0 0.5rem 0; padding-bottom: 6px; border-bottom: 2px solid #e9ecef;
    }
    .stDataFrame { border-radius: 8px; }
    div[data-testid="stSidebarContent"] { background-color: #1a1a2e; }
    div[data-testid="stSidebarContent"] * { color: white !important; }
    .upload-note {
        background: #fff3cd; border-radius: 8px; padding: 0.8rem 1rem;
        font-size: 0.88rem; border-left: 4px solid #ffc107; margin-bottom: 1rem;
    }
</style>
""", unsafe_allow_html=True)

PILLAR_COLORS = {
    'CX Service': '#4f46e5', 'CX Sales': '#7c3aed', 'CX CPQ': '#a855f7',
    'PRC': '#0ea5e9', 'HCM': '#10b981', 'ERP': '#f59e0b',
    'SCM': '#ef4444', 'Test': '#94a3b8', 'ARTestHCM': '#64748b'
}

DATA_FILES = [
    "CX 26C New Features.csv", "PRC 26C New Features.csv",
    "HCM 26C New Features.csv", "ERP 26C New Features.csv", "SCM 26C New Features.csv",
]

def inherit_primary_dates(df):
    fc_col    = 'Feature Classification'
    cfvn_col  = 'Combined Feature Video Name'
    date_cols = ['Actual Release Date', 'Video Ready Date', 'Recording Date']
    if fc_col not in df.columns or cfvn_col not in df.columns:
        return df
    primary_map = {}
    for _, row in df[df[fc_col] == 'Combined Primary NF'].iterrows():
        cfvn = row.get(cfvn_col)
        if pd.notna(cfvn) and str(cfvn).strip():
            primary_map[str(cfvn).strip()] = {col: row.get(col) for col in date_cols}
    if not primary_map:
        return df
    for idx, row in df[df[fc_col] == 'Combined Bundled NF'].iterrows():
        cfvn = row.get(cfvn_col)
        if pd.isna(cfvn) or not str(cfvn).strip():
            continue
        key = str(cfvn).strip()
        if key not in primary_map:
            continue
        for col in date_cols:
            if col in df.columns and pd.isna(row.get(col)) and pd.notna(primary_map[key].get(col)):
                df.at[idx, col] = primary_map[key][col]
    return df

@st.cache_data
def load_data():
    dfs = []
    for f in DATA_FILES:
        try:
            df = pd.read_csv(f)
            dfs.append(df)
        except FileNotFoundError:
            st.error(f"❌ File not found: {f}.")
            st.stop()
    combined = pd.concat(dfs, ignore_index=True)
    combined = combined[~combined['Pillar'].isin(['Test', 'ARTestHCM'])]
    combined = combined[combined['Feature Category'] != 'test']
    valid_from = pd.Timestamp('2020-01-01')
    for col in ['Recording Date', 'Video Ready Date', 'Actual Release Date']:
        if col in combined.columns:
            combined[col] = pd.to_datetime(combined[col], errors='coerce')
            combined[col] = combined[col].apply(
                lambda x: x if pd.notna(x) and x >= valid_from else pd.NaT)
    combined = inherit_primary_dates(combined)
    return combined

def metric_card(value, label):
    return f"""<div class="metric-card"><div class="metric-value">{value}</div><div class="metric-label">{label}</div></div>"""

def export_excel(df):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Report')
    return buf.getvalue()

def color_bar_chart(df, x, y, title, color_col=None):
    fig = px.bar(df, x=x, y=y, title=title, color=color_col or x,
                 color_discrete_map=PILLAR_COLORS, text=y)
    fig.update_traces(textposition='outside')
    fig.update_layout(plot_bgcolor='white', paper_bgcolor='white', font=dict(family='Segoe UI'),
                      showlegend=False, title_font_size=15, margin=dict(t=50, b=40),
                      xaxis_title='', yaxis_title='Count')
    return fig

def add_total_row(df, label_col='Pillar', label='Total'):
    numeric_cols = df.select_dtypes(include='number').columns.tolist()
    total_row = {}
    for col in df.columns:
        if col == label_col: total_row[col] = label
        elif col in numeric_cols: total_row[col] = df[col].sum()
        else: total_row[col] = ''
    return pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

def style_with_total(df, label_col='Pillar', label='Total'):
    def _style(row):
        if row[label_col] == label:
            return ['font-weight: bold; background-color: #f0f0f0'] * len(row)
        return [''] * len(row)
    return df.style.apply(_style, axis=1).hide(axis='index')

def col_config_compact(df):
    config = {}
    for col in df.columns:
        col_len = len(str(col))
        if col_len <= 5: config[col] = st.column_config.Column(width=70)
        elif col_len <= 10: config[col] = st.column_config.Column(width=100)
        elif col_len <= 20: config[col] = st.column_config.Column(width=140)
        elif col_len <= 35: config[col] = st.column_config.Column(width=180)
        else: config[col] = st.column_config.Column(width=250)
    return config

def render_report_filters(df, report_key, has_cdl=True):
    all_pillars = sorted(df['Pillar'].dropna().unique().tolist())
    fc1, fc2 = st.columns(2) if has_cdl else (st.columns(1)[0], None)
    with fc1:
        sel_pillar = st.selectbox("🔍 Filter by Pillar", ['All'] + all_pillars, key=f'pillar_filter_{report_key}')
    sel_cdl = 'All'
    if has_cdl and fc2 is not None:
        if sel_pillar != 'All':
            cdl_options = sorted(df[df['Pillar'] == sel_pillar]['CDL Name'].dropna().unique().tolist())
        else:
            cdl_options = sorted(df['CDL Name'].dropna().unique().tolist())
        cdl_options = [c for c in cdl_options if c not in ['Unassigned', '']]
        with fc2:
            sel_cdl = st.selectbox("🔍 Filter by CDL", ['All'] + cdl_options, key=f'cdl_filter_{report_key}')
    filtered = df.copy()
    if sel_pillar != 'All': filtered = filtered[filtered['Pillar'] == sel_pillar]
    if sel_cdl != 'All' and has_cdl: filtered = filtered[filtered['CDL Name'] == sel_cdl]
    return filtered

with st.sidebar:
    st.markdown("### 📊 26C Features Agent")
    st.markdown("Data auto-loads from the latest CSVs in GitHub.")
    if st.button("🔄 Refresh Data"):
        st.cache_data.clear()
        st.rerun()
    st.markdown("---")
    st.markdown("### 📊 Reports")
    report_options = [
        "🏠 Overview Dashboard",
        "1️⃣  Training Required — Pillar Wise",
        "2️⃣  Training Required — Pillar & Product Wise",
        "3️⃣  AI Tagged Features",
        "4️⃣  DCFR Tagged Features",
        "5️⃣  CDL Coverage — Pillar & CDL Wise",
        "6️⃣  CDL Recording Status — Pillar & CDL Wise",
        "7️⃣  Unboxing Videos — Pillar & Product Wise",
        "8️⃣  Released Videos (incl. Unboxing)",
        "9️⃣  Feature Overall Status",
        "🔟  Released Course List",
        "1️⃣1️⃣  Issues Tracker — Pillar & CDL Wise",
        "1️⃣2️⃣  GA+ Release",
        "1️⃣3️⃣  26C New Features Release Dashboard",
        "1️⃣4️⃣  Daily CDL Status Report",
    ]
    selected_report = st.radio("Select a Report", report_options, label_visibility="collapsed")

st.markdown('<div class="report-title">📊 26C New Features — Report Agent</div>', unsafe_allow_html=True)
st.markdown('<div class="report-subtitle">Pillar-wise reporting across CX, PRC, HCM, ERP, SCM</div>', unsafe_allow_html=True)

df = load_data()
pillars = sorted(df['Pillar'].dropna().unique().tolist())

with st.sidebar:
    st.markdown("---")
    st.markdown("### 🔍 Filter")
    selected_pillars = st.multiselect("Filter by Pillar", pillars, default=pillars)

df_filtered = df[df['Pillar'].isin(selected_pillars)] if selected_pillars else df

# ─────────────────────────────────────────────
# OVERVIEW DASHBOARD
# ─────────────────────────────────────────────
if selected_report == "🏠 Overview Dashboard":
    df_filtered = render_report_filters(df_filtered, 'overview', has_cdl=False)
    df_nf = df_filtered[df_filtered['Feature Category'] != 'Unboxing']
    df_ub = df_filtered[df_filtered['Feature Category'] == 'Unboxing']

    total_nf     = len(df_nf)
    training_yes = len(df_nf[df_nf['Training Required? '] == 'Yes'])
    training_no  = len(df_nf[df_nf['Training Required? '] == 'No'])
    training_tbd = len(df_nf[df_nf['Training Required? '] == 'TBD'])
    rel_nf       = len(df_nf[df_nf['Final Overall Status'] == 'Released (A)'])
    total_ub     = len(df_ub)
    ub_yes       = len(df_ub[df_ub['Training Required? '] == 'Yes'])
    rel_ub       = len(df_ub[df_ub['Final Overall Status'] == 'Released (A)'])
    rel_ub_yes   = len(df_ub[(df_ub['Final Overall Status'] == 'Released (A)') & (df_ub['Training Required? '] == 'Yes')])

    training_blank_df = df_nf[df_nf['Training Required? '].isna()]
    training_blank    = len(training_blank_df)
    pillar_blank      = training_blank_df.groupby('Pillar').size()
    pillar_blank_str  = ', '.join([f"{p}: {c}" for p, c in pillar_blank.items()])

    nf_rel_pct   = round(rel_nf / training_yes * 100, 1) if training_yes > 0 else 0.0
    nf_rel_color = '#10b981' if nf_rel_pct >= 50 else '#ef4444'
    ub_rel_pct   = round(rel_ub_yes / ub_yes * 100, 1) if ub_yes > 0 else 0.0
    ub_rel_color = '#10b981' if ub_rel_pct >= 50 else '#ef4444'

    st.markdown('<div class="section-header">Overall Summary — New Features</div>', unsafe_allow_html=True)
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.markdown(metric_card(total_nf, "Total New Features"), unsafe_allow_html=True)
    c2.markdown(f'<div class="metric-card" style="border-left-color:#4f46e5"><div class="metric-value" style="color:#4f46e5">{training_yes}</div><div class="metric-label">Training = Yes</div></div>', unsafe_allow_html=True)
    c3.markdown(f'<div class="metric-card" style="border-left-color:#10b981"><div class="metric-value" style="color:#10b981">{training_no}</div><div class="metric-label">Training = No</div></div>', unsafe_allow_html=True)
    c4.markdown(f'<div class="metric-card" style="border-left-color:#f59e0b"><div class="metric-value" style="color:#f59e0b">{training_tbd}</div><div class="metric-label">Training = TBD</div></div>', unsafe_allow_html=True)
    c5.markdown(
        f'<div class="metric-card" style="border-left-color:#10b981; text-align:left; padding:1rem;">'
        f'<div style="font-size:1rem; font-weight:700; color:#1a1a2e; margin-bottom:0.4rem;">New Features Released</div>'
        f'<div style="font-size:1.6rem; font-weight:700; color:#10b981;">{rel_nf}</div>'
        f'<div style="font-size:0.85rem; font-weight:700; color:{nf_rel_color}; margin-top:0.3rem;">🚀 Release %: {nf_rel_pct}%</div>'
        f'<div style="font-size:0.75rem; color:#94a3b8;">of Training = Yes ({training_yes})</div>'
        f'</div>', unsafe_allow_html=True)

    if training_blank > 0:
        st.markdown(f'<div class="upload-note">⚠️ <b>{training_blank} features</b> have no Training Required value. ({pillar_blank_str})</div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="section-header">Unboxing Summary</div>', unsafe_allow_html=True)

    # ── CHANGE 1A: 3 columns — Total | Training=Yes | Released ──
    u1, u2, u3 = st.columns(3)
    u1.markdown(metric_card(total_ub, "Total Unboxing"), unsafe_allow_html=True)
    u2.markdown(
        f'<div class="metric-card" style="border-left-color:#4f46e5; text-align:left; padding:1rem;">'
        f'<div style="font-size:1rem; font-weight:700; color:#1a1a2e; margin-bottom:0.4rem;">Unboxing Training = Yes</div>'
        f'<div style="font-size:1.6rem; font-weight:700; color:#4f46e5;">{ub_yes}</div>'
        f'<div style="font-size:0.75rem; color:#94a3b8;">of {total_ub} total Unboxing</div>'
        f'</div>', unsafe_allow_html=True)
    u3.markdown(
        f'<div class="metric-card" style="border-left-color:#7c3aed; text-align:left; padding:1rem;">'
        f'<div style="font-size:1rem; font-weight:700; color:#1a1a2e; margin-bottom:0.4rem;">Unboxing Released</div>'
        f'<div style="font-size:1.6rem; font-weight:700; color:#7c3aed;">{rel_ub}</div>'
        f'<div style="font-size:0.85rem; font-weight:700; color:{ub_rel_color}; margin-top:0.3rem;">🚀 Release %: {ub_rel_pct}%</div>'
        f'<div style="font-size:0.75rem; color:#94a3b8;">of Training = Yes ({ub_yes})</div>'
        f'</div>', unsafe_allow_html=True)

    st.markdown("---")
    st.markdown('<div class="section-header">Pillar Wise — Features & Release Overview</div>', unsafe_allow_html=True)
    pillars_list = sorted(df_filtered['Pillar'].dropna().unique().tolist())
    bar_data = []
    for p in pillars_list:
        p_nf = df_nf[df_nf['Pillar'] == p]
        p_ub = df_ub[df_ub['Pillar'] == p]
        bar_data.append({
            'Pillar': p,
            'Total New Features':    len(p_nf),
            'Training = Yes':        len(p_nf[p_nf['Training Required? '] == 'Yes']),
            'Released New Features': len(p_nf[p_nf['Final Overall Status'] == 'Released (A)']),
            'Total Unboxing':        len(p_ub),
            'Released Unboxing':     len(p_ub[p_ub['Final Overall Status'] == 'Released (A)']),
        })
    bar_df = pd.DataFrame(bar_data)
    bar_melted = bar_df.melt(id_vars='Pillar',
        value_vars=['Total New Features','Training = Yes','Released New Features','Total Unboxing','Released Unboxing'],
        var_name='Metric', value_name='Count')
    COLOR_MAP = {
        'Total New Features': '#4f46e5', 'Training = Yes': '#f59e0b',
        'Released New Features': '#10b981', 'Total Unboxing': '#7c3aed', 'Released Unboxing': '#a855f7',
    }
    fig = px.bar(bar_melted, x='Pillar', y='Count', color='Metric', barmode='group',
                 text='Count', color_discrete_map=COLOR_MAP, title='Pillar Wise — New Features & Unboxing Overview')
    fig.update_traces(textposition='outside')
    fig.update_layout(plot_bgcolor='white', paper_bgcolor='white', font=dict(family='Segoe UI'),
                      title_font_size=15, margin=dict(t=80, b=80), xaxis_title='', yaxis_title='Count',
                      legend=dict(orientation='h', yanchor='top', y=-0.2, xanchor='center', x=0.5, title_text=''))
    st.plotly_chart(fig, use_container_width=True)

# ─────────────────────────────────────────────
# REPORT 1
# ─────────────────────────────────────────────
elif selected_report == "1️⃣  Training Required — Pillar Wise":
    st.markdown('<div class="section-header">Training Required — Pillar Wise (Excluding Unboxing)</div>', unsafe_allow_html=True)
    df_filtered = render_report_filters(df_filtered, 'r1', has_cdl=True)
    df_r1 = df_filtered[df_filtered['Feature Category'] != 'Unboxing'].copy()
    col1, col2, col3 = st.columns(3)
    for col, val_filter, label, color in zip(
        [col1, col2, col3], ['Yes', 'No', 'TBD'],
        ['Training = Yes', 'Training = No', 'Training = TBD'],
        ['#4f46e5', '#10b981', '#f59e0b']
    ):
        count = len(df_r1[df_r1['Training Required? '] == val_filter])
        col.markdown(f'<div class="metric-card" style="border-left-color:{color}"><div class="metric-value" style="color:{color}">{count}</div><div class="metric-label">{label}</div></div>', unsafe_allow_html=True)
    st.markdown("---")
    df_r1['Training Required? '] = df_r1['Training Required? '].fillna('Not Set')
    pivot_r1 = df_r1.groupby(['Pillar', 'Training Required? ']).size().reset_index(name='Count')
    pivot_table = pivot_r1.pivot_table(index='Pillar', columns='Training Required? ', values='Count', fill_value=0)
    for col in ['Yes', 'No', 'TBD', 'Not Set']:
        if col not in pivot_table.columns: pivot_table[col] = 0
    cols_order = [c for c in ['Yes', 'No', 'TBD', 'Not Set'] if c in pivot_table.columns]
    pivot_table = pivot_table[cols_order]
    pivot_table['Total'] = pivot_table.sum(axis=1)
    pivot_table = pivot_table.reset_index()
    fig = px.bar(pivot_r1, x='Pillar', y='Count', color='Training Required? ',
                 barmode='group', title='Training Required by Pillar',
                 color_discrete_map={'Yes':'#4f46e5','No':'#10b981','TBD':'#f59e0b','Not Set':'#94a3b8'}, text='Count')
    fig.update_traces(textposition='outside')
    fig.update_layout(plot_bgcolor='white', paper_bgcolor='white')
    st.plotly_chart(fig, use_container_width=True)
    st.markdown('<div class="section-header">Pillar Wise — Yes / No / TBD Columns</div>', unsafe_allow_html=True)
    pivot_with_total = add_total_row(pivot_table, label_col='Pillar')
    cols_to_fix = ['Yes', 'No', 'TBD', 'Not Set', 'Total']
    existing_cols = [c for c in cols_to_fix if c in pivot_with_total.columns]
    pivot_with_total[existing_cols] = pivot_with_total[existing_cols].astype(int)
    if 'Not Set' in pivot_with_total.columns:
        not_set_data = pivot_with_total[pivot_with_total['Pillar'] != 'Total']['Not Set']
        if (not_set_data == 0).all():
            pivot_with_total = pivot_with_total.drop(columns=['Not Set'])
    styled_pivot = style_with_total(pivot_with_total, label_col='Pillar')
    st.dataframe(styled_pivot, use_container_width=True, column_config=col_config_compact(pivot_with_total))
    st.markdown("---")
    st.markdown('<div class="section-header">Feature Detail — Training = Yes</div>', unsafe_allow_html=True)
    tr_yes = df_r1[df_r1['Training Required? '] == 'Yes'][['Pillar','Product','Module','Feature','CDL Name','Feature Category']].reset_index(drop=True)
    st.dataframe(tr_yes, use_container_width=True, column_config=col_config_compact(tr_yes))
    st.download_button("⬇️ Download Report", export_excel(tr_yes), "report1_training_yes.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    df_not_set = df_r1[df_r1['Training Required? '] == 'Not Set'][['Pillar', 'Product', 'CDL Name']].copy()
    df_not_set['CDL Name'] = df_not_set['CDL Name'].fillna('Unassigned')
    if len(df_not_set) > 0:
        pillar_summary = df_not_set.groupby('Pillar').size()
        pillar_summary_str = ', '.join([f"{p}: {c}" for p, c in pillar_summary.items()])
        subpoints_html = ''
        for pillar, pillar_df in df_not_set.groupby('Pillar'):
            subpoints_html += f'<li style="margin-bottom:6px;"><b style="color:#1a1a2e;">{pillar}</b><ul style="margin-top:4px;">'
            for product, prod_df in pillar_df.groupby('Product'):
                cdl_names = ', '.join(sorted(prod_df['CDL Name'].unique().tolist()))
                subpoints_html += f'<li style="margin-bottom:3px;"><b>{product}</b> — CDL: {cdl_names}</li>'
            subpoints_html += '</ul></li>'
        st.markdown(f"""<div style="background-color:#fffbeb;padding:12px 16px;border-radius:6px;border-left:4px solid #f59e0b;margin-top:8px;font-size:0.82rem;color:#444;">
            <b>⚠️ {len(df_not_set)} features have no Training Required value. ({pillar_summary_str})</b><br><br>
            The following features are missing the <b>Training Required</b> field:<ul style="margin-top:8px;line-height:1.8;">{subpoints_html}</ul>
        </div>""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# REPORT 2
# ─────────────────────────────────────────────
elif selected_report == "2️⃣  Training Required — Pillar & Product Wise":
    st.markdown('<div class="section-header">Training Required — Pillar & Product Wise</div>', unsafe_allow_html=True)
    df_filtered = render_report_filters(df_filtered, 'r2', has_cdl=True)
    df_r2 = df_filtered[df_filtered['Feature Category'] != 'Unboxing'].copy()
    df_r2_ub = df_filtered[df_filtered['Feature Category'] == 'Unboxing'].copy()
    df_r2['Training Required? '] = df_r2['Training Required? '].fillna('Not Set')
    df_r2_ub['Training Required? '] = df_r2_ub['Training Required? '].fillna('Not Set')
    def map_pillar_group(pillar):
        if str(pillar).startswith('CX'): return 'CX'
        return pillar
    df_r2['Pillar Group']    = df_r2['Pillar'].apply(map_pillar_group)
    df_r2_ub['Pillar Group'] = df_r2_ub['Pillar'].apply(map_pillar_group)
    st.markdown('<div class="section-header">📦 Pillar Wise — Training Summary</div>', unsafe_allow_html=True)
    pillar_groups = ['CX', 'ERP', 'HCM', 'PRC', 'SCM']
    box_cols = st.columns(5)
    for col, pg in zip(box_cols, pillar_groups):
        pg_df    = df_r2[df_r2['Pillar Group'] == pg]
        pg_ub_df = df_r2_ub[df_r2_ub['Pillar Group'] == pg]
        total   = len(pg_df)
        yes     = len(pg_df[pg_df['Training Required? '] == 'Yes'])
        no      = len(pg_df[pg_df['Training Required? '] == 'No'])
        tbd     = len(pg_df[pg_df['Training Required? '] == 'TBD'])
        not_set = len(pg_df[pg_df['Training Required? '] == 'Not Set'])
        coverage_pct   = round(yes / total * 100, 1) if total > 0 else 0.0
        coverage_color = '#10b981' if coverage_pct >= 50 else '#ef4444'
        nf_released  = len(pg_df[(pg_df['Training Required? '] == 'Yes') & (pg_df['Final Overall Status'] == 'Released (A)')])
        nf_rel_pct   = round(nf_released / yes * 100, 1) if yes > 0 else 0.0
        nf_rel_color = '#10b981' if nf_rel_pct >= 50 else '#ef4444'
        ub_yes_count = len(pg_ub_df[pg_ub_df['Training Required? '] == 'Yes'])
        ub_released  = len(pg_ub_df[(pg_ub_df['Training Required? '] == 'Yes') & (pg_ub_df['Final Overall Status'] == 'Released (A)')])
        ub_rel_pct   = round(ub_released / ub_yes_count * 100, 1) if ub_yes_count > 0 else 0.0
        ub_rel_color = '#10b981' if ub_rel_pct >= 50 else '#ef4444'
        col.markdown(
            f'<div class="metric-card" style="border-left-color:#4f46e5; text-align:left; padding:1rem;">'
            f'<div style="font-size:1rem; font-weight:700; color:#1a1a2e; margin-bottom:0.5rem;">{pg}</div>'
            f'<div style="font-size:0.85rem; color:#4f46e5;">Yes: <b>{yes}</b></div>'
            f'<div style="font-size:0.85rem; color:#10b981;">No: <b>{no}</b></div>'
            f'<div style="font-size:0.85rem; color:#f59e0b;">TBD: <b>{tbd}</b></div>'
            f'<div style="font-size:0.85rem; color:#94a3b8;">Not Set: <b>{not_set}</b></div>'
            f'<div style="font-size:0.9rem; font-weight:700; color:{coverage_color}; margin-top:0.4rem;">% Coverage: {coverage_pct}%</div>'
            f'<div style="border-top:1px solid #e9ecef; margin:0.4rem 0;"></div>'
            f'<div style="font-size:0.82rem; font-weight:700; color:{nf_rel_color};">🚀 NF Release %: {nf_rel_pct}%</div>'
            f'<div style="font-size:0.75rem; color:#94a3b8; margin-bottom:0.2rem;">({nf_released} of {yes} Training=Yes NF)</div>'
            f'<div style="font-size:0.82rem; font-weight:700; color:{ub_rel_color};">📦 Unboxing Release %: {ub_rel_pct}%</div>'
            f'<div style="font-size:0.75rem; color:#94a3b8;">({ub_released} of {ub_yes_count} Training=Yes Unboxing)</div>'
            f'</div>', unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown('<div class="section-header">📊 Training Required Summary — Pillar & Product Wise</div>', unsafe_allow_html=True)
    pivot = df_r2.groupby(['Pillar', 'Product', 'Training Required? ']).size().unstack(fill_value=0)
    for c in ['Yes', 'No', 'TBD', 'Not Set']:
        if c not in pivot.columns: pivot[c] = 0
    pivot = pivot[['Yes', 'No', 'TBD', 'Not Set']]
    pivot['Total'] = pivot.sum(axis=1)
    pivot['Yes %'] = pivot.apply(lambda row: round((row['Yes'] / row['Total']) * 100, 1) if row['Total'] > 0 else 0.0, axis=1)
    pivot = pivot.reset_index()
    pivot_with_total = add_total_row(pivot, label_col='Pillar')
    total_mask = pivot_with_total['Pillar'] == 'Total'
    yes_sum = pivot_with_total.loc[total_mask, 'Yes'].values[0]
    grand_total = pivot_with_total.loc[total_mask, 'Total'].values[0]
    if grand_total > 0:
        pivot_with_total.loc[total_mask, 'Yes %'] = round((yes_sum / grand_total) * 100, 1)
    cols_to_fix = ['Yes', 'No', 'TBD', 'Not Set', 'Total']
    pivot_with_total[cols_to_fix] = pivot_with_total[cols_to_fix].astype(int)
    def style_rows(row):
        if row['Pillar'] == 'Total': return ['font-weight: bold; background-color: #f0f0f0'] * len(row)
        yes_val = row['Yes']; yes_pct = row['Yes %']
        if yes_val == 0: return ['color: red; font-weight: bold'] * len(row)
        elif isinstance(yes_pct, (int, float)) and yes_pct >= 1.0 and yes_pct < 50.0: return ['color: darkorange; font-weight: bold'] * len(row)
        return [''] * len(row)
    styled_pivot = pivot_with_total.style.apply(style_rows, axis=1).format({'Yes %': lambda x: f'{x:.1f}%' if isinstance(x, (int, float)) else x}).hide(axis='index')
    st.dataframe(styled_pivot, use_container_width=True, column_config=col_config_compact(pivot_with_total))
    st.download_button("⬇️ Download Training Summary", export_excel(pivot), "report2_training_summary.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_r2_summary")
    st.markdown("---")
    st.markdown('<div class="section-header">📋 Features — Training = No / TBD / Not Set (with Reason)</div>', unsafe_allow_html=True)
    df_r2_detail = df_r2[df_r2['Training Required? '].isin(['No', 'TBD', 'Not Set'])].copy()
    reason_col = 'Reason why NF content cannot be built?'
    cols_to_show = ['Pillar', 'Product', 'Module', 'Feature', 'CDL Name', 'Training Required? ', reason_col]
    cols_available = [c for c in cols_to_show if c in df_r2_detail.columns]
    detail_table = df_r2_detail[cols_available].reset_index(drop=True)
    def style_table2(row):
        training_val = row.get('Training Required? ', ''); reason_val = row.get(reason_col, '')
        reason_blank = pd.isna(reason_val) or str(reason_val).strip() == ''
        if training_val == 'No' and reason_blank: return ['color: red; font-weight: bold'] * len(row)
        elif training_val == 'TBD' and reason_blank: return ['color: darkorange; font-weight: bold'] * len(row)
        return [''] * len(row)
    if reason_col in detail_table.columns:
        st.dataframe(detail_table.style.apply(style_table2, axis=1).hide(axis='index'), use_container_width=True, column_config=col_config_compact(detail_table))
    else:
        st.dataframe(detail_table, use_container_width=True, column_config=col_config_compact(detail_table))
    st.markdown("""<div style="background-color:#fff8f8;padding:10px 14px;border-radius:6px;border-left:4px solid #ef4444;margin-top:8px;font-size:0.82rem;color:#444;">
        <b>📌 Row Highlight Guide:</b><br>
        <span style="color:red;font-weight:bold;">🔴 Red rows</span> — Training = <b>No</b> but no reason provided.<br>
        <span style="color:darkorange;font-weight:bold;">🟠 Orange rows</span> — Training = <b>TBD</b> but no reason provided.
    </div>""", unsafe_allow_html=True)
    st.download_button("⬇️ Download No / TBD / Not Set Features", export_excel(detail_table), "report2_no_tbd_notset.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_r2_detail")

# ─────────────────────────────────────────────
# REPORT 3: AI Tagged Features
# ─────────────────────────────────────────────
elif selected_report == "3️⃣  AI Tagged Features":
    st.markdown('<div class="section-header">AI Tagged Features Across All Pillars</div>', unsafe_allow_html=True)
    df_filtered = render_report_filters(df_filtered, 'r3', has_cdl=True)
    st.info("ℹ️ Any feature with 'AI' anywhere in its tag is counted as an AI feature. 'Agentic App' features are tracked separately.")
    df_r3 = df_filtered[df_filtered['Feature Category'] != 'Unboxing'].copy()
    df_agentic = df_r3[df_r3['Tags ( AI / Redwood/AI Agent)'].str.contains('agentic app', case=False, na=False)].copy()
    df_ai = df_r3[df_r3['Tags ( AI / Redwood/AI Agent)'].str.contains('AI', na=False) & ~df_r3['Tags ( AI / Redwood/AI Agent)'].str.contains('agentic app', case=False, na=False)].copy()
    col1, col2, col4, col3 = st.columns(4)
    ai_yes = len(df_ai[df_ai['Training Required? '] == 'Yes'])
    ai_rel = len(df_ai[df_ai['Final Overall Status'] == 'Released (A)'])
    ai_yes_pct = round(ai_yes / len(df_ai) * 100, 1) if len(df_ai) > 0 else 0
    ai_rel_pct = round(ai_rel / ai_yes * 100, 1) if ai_yes > 0 else 0
    col1.markdown(f"""<div class="metric-card" style="text-align:left; padding:1rem;">
        <div style="font-size:1.6rem; font-weight:700; color:#4f46e5; margin-bottom:0.3rem;">{len(df_ai)}</div>
        <div style="font-size:0.85rem; color:#6c757d; margin-bottom:0.6rem; font-weight:600;">AI Tagged Features</div>
        <div style="font-size:0.8rem; color:#1a1a2e; margin-bottom:0.2rem;">📌 Identification: <b>{len(df_ai)}</b> AI features</div>
        <div style="font-size:0.8rem; color:#4f46e5; margin-bottom:0.2rem;">🎯 Training: <b>{ai_yes}</b> <span style="color:#94a3b8;">({ai_yes_pct}%)</span></div>
        <div style="font-size:0.8rem; color:#10b981;">🚀 Released: <b>{ai_rel}</b> <span style="color:#94a3b8;">({ai_rel_pct}% of Training=Yes)</span></div>
    </div>""", unsafe_allow_html=True)
    ag_yes = len(df_agentic[df_agentic['Training Required? '] == 'Yes'])
    ag_rel = len(df_agentic[df_agentic['Final Overall Status'] == 'Released (A)'])
    ag_yes_pct = round(ag_yes / len(df_agentic) * 100, 1) if len(df_agentic) > 0 else 0
    ag_rel_pct = round(ag_rel / ag_yes * 100, 1) if ag_yes > 0 else 0
    col2.markdown(f"""<div class="metric-card" style="text-align:left; padding:1rem;">
        <div style="font-size:1.6rem; font-weight:700; color:#7c3aed; margin-bottom:0.3rem;">{len(df_agentic)}</div>
        <div style="font-size:0.85rem; color:#6c757d; margin-bottom:0.6rem; font-weight:600;">AI Agentic App Features</div>
        <div style="font-size:0.8rem; color:#1a1a2e; margin-bottom:0.2rem;">📌 Identification: <b>{len(df_agentic)}</b> Agentic features</div>
        <div style="font-size:0.8rem; color:#7c3aed; margin-bottom:0.2rem;">🎯 Training: <b>{ag_yes}</b> <span style="color:#94a3b8;">({ag_yes_pct}%)</span></div>
        <div style="font-size:0.8rem; color:#10b981;">🚀 Released: <b>{ag_rel}</b> <span style="color:#94a3b8;">({ag_rel_pct}% of Training=Yes)</span></div>
    </div>""", unsafe_allow_html=True)
    col4.markdown(f"""<div class="metric-card" style="text-align:left; padding:1rem; border-left-color:#10b981;">
        <div style="font-size:0.95rem; font-weight:700; color:#10b981; margin-bottom:0.4rem;">Training Coverage</div>
        <div style="font-size:0.78rem; color:#4f46e5; font-weight:700; margin-bottom:0.2rem; margin-top:0.4rem;">🤖 AI Tagged ({len(df_ai)} features)</div>
        <div style="font-size:0.78rem; color:#1a1a2e; margin-bottom:0.1rem;">🎯 Training = Yes: <b>{ai_yes}</b> ({ai_yes_pct}%)</div>
        <div style="font-size:0.78rem; color:#10b981; margin-bottom:0.4rem;">🚀 Released: <b>{ai_rel}</b> ({ai_rel_pct}%)</div>
        <div style="border-top:1px solid #e9ecef; margin: 0.4rem 0;"></div>
        <div style="font-size:0.78rem; color:#7c3aed; font-weight:700; margin-bottom:0.2rem;">🧠 AI Agentic App ({len(df_agentic)} features)</div>
        <div style="font-size:0.78rem; color:#1a1a2e; margin-bottom:0.1rem;">🎯 Training = Yes: <b>{ag_yes}</b> ({ag_yes_pct}%)</div>
        <div style="font-size:0.78rem; color:#10b981;">🚀 Released: <b>{ag_rel}</b> ({ag_rel_pct}%)</div>
    </div>""", unsafe_allow_html=True)
    ai_pct = round(len(df_ai) / len(df_r3) * 100, 1) if len(df_r3) > 0 else 0
    ag_pct = round(len(df_agentic) / len(df_r3) * 100, 1) if len(df_r3) > 0 else 0
    col3.markdown(f"""<div class="metric-card" style="text-align:left; padding:1rem;">
        <div style="font-size:1.6rem; font-weight:700; color:#4f46e5; margin-bottom:0.3rem;">{ai_pct}%</div>
        <div style="font-size:0.85rem; color:#6c757d; margin-bottom:0.6rem; font-weight:600;">AI % of Total New Features</div>
        <div style="font-size:0.8rem; color:#4f46e5; margin-bottom:0.3rem;">🤖 AI Tagged: <b>{len(df_ai)}</b> / {len(df_r3)} = <b>{ai_pct}%</b></div>
        <div style="font-size:0.8rem; color:#7c3aed;">🧠 AI Agentic App: <b>{len(df_agentic)}</b> / {len(df_r3)} = <b>{ag_pct}%</b></div>
    </div>""", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown('<div class="section-header">Pillar Wise — AI Feature Summary</div>', unsafe_allow_html=True)
    pillar_total   = df_r3.groupby('Pillar').size().reset_index(name='Total NF')
    pillar_ai      = df_ai.groupby('Pillar').size().reset_index(name='AI Features')
    pillar_agentic = df_agentic.groupby('Pillar').size().reset_index(name='AI Agentic App')
    pillar_ai_training = df_ai[df_ai['Training Required? '] == 'Yes'].groupby('Pillar').size().reset_index(name='AI Training Req')
    pillar_ag_training = df_agentic[df_agentic['Training Required? '] == 'Yes'].groupby('Pillar').size().reset_index(name='Agentic Training Req')
    pillar_ai_released = df_ai[df_ai['Final Overall Status'] == 'Released (A)'].groupby('Pillar').size().reset_index(name='AI Released')
    pillar_ag_released = df_agentic[df_agentic['Final Overall Status'] == 'Released (A)'].groupby('Pillar').size().reset_index(name='Agentic Released')
    ai_summary = pillar_total.merge(pillar_ai, on='Pillar', how='left').merge(pillar_agentic, on='Pillar', how='left').merge(pillar_ai_training, on='Pillar', how='left').merge(pillar_ag_training, on='Pillar', how='left').merge(pillar_ai_released, on='Pillar', how='left').merge(pillar_ag_released, on='Pillar', how='left').fillna(0)
    for c in ['AI Features','AI Agentic App','AI Training Req','Agentic Training Req','AI Released','Agentic Released']:
        ai_summary[c] = ai_summary[c].astype(int)
    ai_summary['AI %'] = ai_summary.apply(lambda r: round(r['AI Training Req']/r['AI Features']*100,1) if r['AI Features']>0 else 0.0, axis=1)
    ai_summary['AI Agentic %'] = ai_summary.apply(lambda r: round(r['Agentic Training Req']/r['AI Agentic App']*100,1) if r['AI Agentic App']>0 else 0.0, axis=1)
    ai_summary['AI Release %'] = ai_summary.apply(lambda r: round(r['AI Released']/r['AI Training Req']*100,1) if r['AI Training Req']>0 else 0.0, axis=1)
    ai_summary['Agentic Release %'] = ai_summary.apply(lambda r: round(r['Agentic Released']/r['Agentic Training Req']*100,1) if r['Agentic Training Req']>0 else 0.0, axis=1)
    ai_summary = ai_summary[['Pillar','Total NF','AI Features','AI Training Req','AI %','AI Agentic App','Agentic Training Req','AI Agentic %','AI Released','AI Release %','Agentic Released','Agentic Release %']]
    ai_summary_with_total = add_total_row(ai_summary)
    total_mask = ai_summary_with_total['Pillar'] == 'Total'
    t = ai_summary_with_total.loc[total_mask].iloc[0]
    ai_summary_with_total.loc[total_mask, 'AI %'] = round(t['AI Training Req']/t['AI Features']*100,1) if t['AI Features']>0 else 0.0
    ai_summary_with_total.loc[total_mask, 'AI Agentic %'] = round(t['Agentic Training Req']/t['AI Agentic App']*100,1) if t['AI Agentic App']>0 else 0.0
    ai_summary_with_total.loc[total_mask, 'AI Release %'] = round(t['AI Released']/t['AI Training Req']*100,1) if t['AI Training Req']>0 else 0.0
    ai_summary_with_total.loc[total_mask, 'Agentic Release %'] = round(t['Agentic Released']/t['Agentic Training Req']*100,1) if t['Agentic Training Req']>0 else 0.0
    count_cols = ['Total NF','AI Features','AI Agentic App','AI Training Req','Agentic Training Req','AI Released','Agentic Released']
    ai_summary_with_total[count_cols] = ai_summary_with_total[count_cols].astype(int)
    pct_fmt = lambda x: f'{x:.1f}%' if isinstance(x, (int, float)) else x
    def style_ai_summary(row):
        if row['Pillar'] == 'Total': return ['font-weight: bold; background-color: #f0f0f0'] * len(row)
        if isinstance(row['AI Release %'], (int, float)) and row['AI Release %'] < 50.0: return ['color: red; font-weight: bold'] * len(row)
        return [''] * len(row)
    styled_ai_summary = ai_summary_with_total.style.apply(style_ai_summary, axis=1).format({'AI %':pct_fmt,'AI Agentic %':pct_fmt,'AI Release %':pct_fmt,'Agentic Release %':pct_fmt}).hide(axis='index')
    st.dataframe(styled_ai_summary, use_container_width=True, column_config=col_config_compact(ai_summary_with_total))
    t_row = ai_summary_with_total.loc[ai_summary_with_total['Pillar'] == 'Total'].iloc[0]
    st.markdown(f"""> 📌 **Notes:**
> - **AI %** = AI Features requiring CDL training ÷ Total AI-tagged Features. *(Currently **{t_row['AI %']}%** overall)*
> - **AI Agentic %** = Agentic Apps requiring CDL training ÷ Total Agentic Apps. *(Currently **{t_row['AI Agentic %']}%** overall)*
> - **AI Release %** = Released AI Features ÷ AI Features (Training Required). *(Currently **{t_row['AI Release %']}%** overall)*
> - **Agentic Release %** = Released Agentic Apps ÷ Agentic Apps (Training Required). *(Currently **{t_row['Agentic Release %']}%** overall)*""")
    st.markdown("---")
    st.markdown('<div class="section-header">AI Tag Types Distribution — Pillar Wise</div>', unsafe_allow_html=True)
    df_ai_chart = pd.concat([df_ai, df_agentic], ignore_index=True)
    tag_pillar  = df_ai_chart.groupby(['Tags ( AI / Redwood/AI Agent)', 'Pillar']).size().reset_index(name='Count')
    tag_totals  = tag_pillar.groupby('Tags ( AI / Redwood/AI Agent)')['Count'].sum().reset_index(name='Total')
    tag_pillar  = tag_pillar.merge(tag_totals, on='Tags ( AI / Redwood/AI Agent)')
    fig_tag = px.bar(tag_pillar, x='Tags ( AI / Redwood/AI Agent)', y='Count', color='Pillar', barmode='stack', title='AI Tag Types Distribution by Pillar', color_discrete_map=PILLAR_COLORS, text='Count')
    fig_tag.update_traces(textposition='inside', textfont_size=11)
    for tag, total in zip(tag_totals['Tags ( AI / Redwood/AI Agent)'], tag_totals['Total']):
        fig_tag.add_annotation(x=tag, y=total, text=f"<b>{total}</b>", showarrow=False, yshift=10, font=dict(size=13, color='black'))
    fig_tag.update_layout(plot_bgcolor='white', paper_bgcolor='white', font=dict(family='Segoe UI'), title_font_size=15, margin=dict(t=60, b=120), xaxis_title='', yaxis_title='Count', xaxis_tickangle=-30, legend=dict(orientation='h', yanchor='top', y=-0.3, xanchor='center', x=0.5, title_text=''))
    st.plotly_chart(fig_tag, use_container_width=True)
    st.markdown("---")
    st.markdown('<div class="section-header">AI Features — Training Required (Pillar Wise)</div>', unsafe_allow_html=True)
    ai_tr = df_ai.copy()
    ai_tr['Training Required? '] = ai_tr['Training Required? '].fillna('Not Set')
    ai_tr_pivot = ai_tr.groupby(['Pillar', 'Training Required? ']).size().reset_index(name='Count')
    fig3 = px.bar(ai_tr_pivot, x='Pillar', y='Count', color='Training Required? ', barmode='group', title='AI Features — Training Required by Pillar', color_discrete_map={'Yes':'#4f46e5','No':'#10b981','TBD':'#f59e0b','Not Set':'#94a3b8'}, text='Count')
    fig3.update_traces(textposition='outside')
    fig3.update_layout(plot_bgcolor='white', paper_bgcolor='white')
    st.plotly_chart(fig3, use_container_width=True)
    ai_tr_table = ai_tr_pivot.pivot_table(index='Pillar', columns='Training Required? ', values='Count', fill_value=0).reset_index()
    ai_tr_table['Total'] = ai_tr_table.select_dtypes(include='number').sum(axis=1)
    ai_tr_with_total = add_total_row(ai_tr_table)
    cols_to_fix = [c for c in ai_tr_with_total.columns if c != 'Pillar']
    ai_tr_with_total[cols_to_fix] = ai_tr_with_total[cols_to_fix].astype(int)
    st.dataframe(style_with_total(ai_tr_with_total), use_container_width=True, column_config=col_config_compact(ai_tr_with_total))
    st.markdown("---")
    st.markdown('<div class="section-header">AI Features — Released Status (Pillar Wise)</div>', unsafe_allow_html=True)
    ai_rel_df = df_ai.copy()
    ai_rel_df['Final Overall Status'] = ai_rel_df['Final Overall Status'].fillna('Not Set')
    ai_rel_pivot = ai_rel_df.groupby(['Pillar', 'Final Overall Status']).size().reset_index(name='Count')
    fig4 = px.bar(ai_rel_pivot, x='Pillar', y='Count', color='Final Overall Status', barmode='group', title='AI Features — Release Status by Pillar', color_discrete_map={'Released (A)':'#10b981','In Review (A)':'#4f46e5','Not Set':'#94a3b8'}, text='Count')
    fig4.update_traces(textposition='outside')
    fig4.update_layout(plot_bgcolor='white', paper_bgcolor='white')
    st.plotly_chart(fig4, use_container_width=True)
    ai_rel_table = ai_rel_pivot.pivot_table(index='Pillar', columns='Final Overall Status', values='Count', fill_value=0).reset_index()
    ai_rel_table['Total'] = ai_rel_table.select_dtypes(include='number').sum(axis=1)
    ai_rel_with_total = add_total_row(ai_rel_table)
    cols_to_fix = [c for c in ai_rel_with_total.columns if c != 'Pillar']
    ai_rel_with_total[cols_to_fix] = ai_rel_with_total[cols_to_fix].astype(int)
    st.dataframe(style_with_total(ai_rel_with_total), use_container_width=True, column_config=col_config_compact(ai_rel_with_total))
    st.markdown("---")
    st.markdown('<div class="section-header">AI Feature Details</div>', unsafe_allow_html=True)
    reason_col = 'Reason why NF content cannot be built?'
    detail_cols = ['Pillar','Product','Module','Feature','CDL Name','Tags ( AI / Redwood/AI Agent)','Training Required? ','Final Overall Status',reason_col]
    detail_cols_available = [c for c in detail_cols if c in df_ai.columns]
    detail = df_ai[detail_cols_available].reset_index(drop=True)
    def style_ai_detail(row):
        training_val = row.get('Training Required? ', ''); reason_val = row.get(reason_col, '')
        reason_blank = pd.isna(reason_val) or str(reason_val).strip() == ''
        if training_val == 'No' and reason_blank: return ['color: red; font-weight: bold'] * len(row)
        elif training_val == 'TBD' and reason_blank: return ['color: darkorange; font-weight: bold'] * len(row)
        return [''] * len(row)
    if reason_col in detail.columns:
        st.dataframe(detail.style.apply(style_ai_detail, axis=1).hide(axis='index'), use_container_width=True, column_config=col_config_compact(detail))
    else:
        st.dataframe(detail, use_container_width=True, column_config=col_config_compact(detail))
    st.download_button("⬇️ Download AI Features Report", export_excel(detail), "report3_ai_features.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    st.markdown("---")
    st.markdown('<div class="section-header">AI Features Released — Pillar Wise by GA+ Tier</div>', unsafe_allow_html=True)

    _GA_BASE_R3 = pd.Timestamp('2026-07-03')

    def get_ga_tier_r3(date, base):
        if pd.isna(date): return 'Not Set'
        if date <= base: return 'GA+0'
        diff = (date - base).days
        tier = ((diff - 1) // 7) + 1
        return f'GA+{tier}'

    df_ai_released = df_ai[df_ai['Final Overall Status'] == 'Released (A)'].copy()

    valid_from_r3 = pd.Timestamp('2020-01-01')
    for col in ['Actual Release Date']:
        if col in df_ai_released.columns:
            df_ai_released[col] = pd.to_datetime(df_ai_released[col], errors='coerce').apply(
                lambda x: x if pd.notna(x) and x >= valid_from_r3 else pd.NaT
            )

    if len(df_ai_released) == 0:
        st.info("No AI features released yet.")
    else:
        df_ai_released['GA Tier'] = df_ai_released['Actual Release Date'].apply(
            lambda x: get_ga_tier_r3(x, _GA_BASE_R3)
        )

        # Get all unique GA tiers sorted
        all_ga_tiers = sorted(
            [t for t in df_ai_released['GA Tier'].unique() if t != 'Not Set'],
            key=lambda t: int(t.replace('GA+', '')) if t.replace('GA+', '').isdigit() else 999
        )
        if 'Not Set' in df_ai_released['GA Tier'].unique():
            all_ga_tiers.append('Not Set')

        # Build pillar-wise pivot
        ga_pivot = df_ai_released.groupby(['Pillar', 'GA Tier']).size().unstack(fill_value=0)
        for tier in all_ga_tiers:
            if tier not in ga_pivot.columns:
                ga_pivot[tier] = 0
        ga_pivot = ga_pivot[all_ga_tiers]
        ga_pivot['Total Released'] = ga_pivot.sum(axis=1)
        ga_pivot = ga_pivot.reset_index()

        ga_pivot_with_total = add_total_row(ga_pivot, label_col='Pillar')
        num_cols = [c for c in ga_pivot_with_total.columns if c != 'Pillar']
        ga_pivot_with_total[num_cols] = ga_pivot_with_total[num_cols].astype(int)

        def style_ga_r3(row):
            if row['Pillar'] == 'Total':
                return ['font-weight: bold; background-color: #f0f0f0'] * len(row)
            styles = []
            for c in list(row.index):
                if c == 'Total Released':
                    styles.append('color: #10b981; font-weight: bold')
                elif c == 'Pillar':
                    styles.append('font-weight: bold')
                elif c != 'Not Set':
                    styles.append('color: #4f46e5; font-weight: bold')
                else:
                    styles.append('color: #94a3b8')
            return styles

        st.dataframe(
            ga_pivot_with_total.style.apply(style_ga_r3, axis=1).hide(axis='index'),
            use_container_width=True,
            column_config=col_config_compact(ga_pivot_with_total)
        )
        st.markdown(
            '<div style="background-color:#e8f4f8; padding:8px 14px; border-radius:6px; '
            'border-left:4px solid #0078d4; margin-top:6px; font-size:0.82rem; color:#444;">'
            '📌 <b>GA+ Tier</b> = derived from <b>Actual Release Date</b> against GA+0 base date '
            f'<b>{_GA_BASE_R3.strftime("%d-%b-%Y")}</b>. Only AI-tagged features with '
            '<b>Final Overall Status = Released (A)</b> are counted here.'
            '</div>',
            unsafe_allow_html=True
        )
        st.download_button(
            "⬇️ Download AI Released by GA+ Tier",
            export_excel(ga_pivot_with_total),
            "report3_ai_released_ga.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_r3_ga"
        )

# ─────────────────────────────────────────────
# REPORT 4: DCFR Tagged Features
# ─────────────────────────────────────────────
elif selected_report == "4️⃣  DCFR Tagged Features":
    st.markdown('<div class="section-header">DCFR Tagged Features — Pillar & Product Wise</div>', unsafe_allow_html=True)
    df_filtered = render_report_filters(df_filtered, 'r4', has_cdl=True)
    df_r4_all = df_filtered[df_filtered['Feature Category'] != 'Unboxing'].copy()
    def map_pillar_group_dcfr(pillar):
        if str(pillar).startswith('CX'): return 'CX'
        return pillar
    df_r4_all['Pillar Group'] = df_r4_all['Pillar'].apply(map_pillar_group_dcfr)
    df_dcfr = df_r4_all[df_r4_all['DCFR Ranking'].notna()].copy()
    col_m1, col_m2, col_m3 = st.columns(3)
    col_m1.markdown(metric_card(len(df_dcfr), "Total DCFR Tagged"), unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="section-header">📦 Pillar Wise — DCFR Summary</div>', unsafe_allow_html=True)
    pillar_groups = ['CX', 'ERP', 'HCM', 'PRC', 'SCM']
    box_cols = st.columns(5)
    for col, pg in zip(box_cols, pillar_groups):
        pg_dcfr = df_dcfr[df_dcfr['Pillar Group'] == pg]
        total   = len(pg_dcfr)
        hl_df   = pg_dcfr[pg_dcfr['DCFR Ranking'].isin(['High', 'Larger Scale'])]
        high_large = len(hl_df)
        hl_train   = len(hl_df[hl_df['Training Required? '] == 'Yes'])
        hl_train_pct = round(hl_train / high_large * 100, 1) if high_large > 0 else 0.0
        hl_released  = len(hl_df[(hl_df['Training Required? '] == 'Yes') & (hl_df['Final Overall Status'] == 'Released (A)')])
        hl_rel_pct   = round(hl_released / hl_train * 100, 1) if hl_train > 0 else 0.0
        low_small = len(pg_dcfr[pg_dcfr['DCFR Ranking'].isin(['Low', 'Small scale'])])
        medium    = len(pg_dcfr[pg_dcfr['DCFR Ranking'] == 'Medium'])
        col.markdown(
            f'<div class="metric-card" style="border-left-color:#4f46e5; text-align:left; padding:1rem;">'
            f'<div style="font-size:1rem; font-weight:700; color:#1a1a2e; margin-bottom:0.5rem;">{pg}</div>'
            f'<div style="font-size:0.85rem; color:#6c757d;">Total DCFR: <b>{total}</b></div>'
            f'<div style="font-size:0.85rem; color:#ef4444; margin-top:0.3rem;">High + Large Scale: <b>{high_large}</b></div>'
            f'<div style="font-size:0.8rem; color:#4f46e5; margin-left:0.8rem;">🎯 Training %: <b>{hl_train_pct}%</b> ({hl_train} of {high_large})</div>'
            f'<div style="font-size:0.8rem; color:#10b981; margin-left:0.8rem;">🚀 Release %: <b>{hl_rel_pct}%</b> ({hl_released} of {hl_train})</div>'
            f'<div style="font-size:0.85rem; color:#10b981; margin-top:0.3rem;">Low + Small Scale: <b>{low_small}</b></div>'
            f'<div style="font-size:0.85rem; color:#f59e0b;">Medium: <b>{medium}</b></div>'
            f'</div>', unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown('<div class="section-header">Pillar Wise — DCFR Ranking Summary</div>', unsafe_allow_html=True)
    all_pillars   = sorted(df_filtered['Pillar'].dropna().unique().tolist())
    dcfr_rankings = ['High', 'Larger Scale', 'Low', 'Medium', 'Small scale']
    dcfr_pivot = df_dcfr.groupby(['Pillar', 'DCFR Ranking']).size().unstack(fill_value=0)
    dcfr_pivot = dcfr_pivot.reindex(all_pillars, fill_value=0)
    for r in dcfr_rankings:
        if r not in dcfr_pivot.columns: dcfr_pivot[r] = 0
    dcfr_pivot = dcfr_pivot[dcfr_rankings]
    dcfr_pivot['Total'] = dcfr_pivot.sum(axis=1)
    dcfr_pivot = dcfr_pivot.reset_index()
    dcfr_pivot_with_total = add_total_row(dcfr_pivot)
    st.dataframe(style_with_total(dcfr_pivot_with_total), use_container_width=True, column_config=col_config_compact(dcfr_pivot_with_total))
    st.markdown("---")
    st.markdown('<div class="section-header">Pillar + Product Wise — DCFR Ranking (with Released)</div>', unsafe_allow_html=True)
    dcfr_prod_pivot = df_dcfr.groupby(['Pillar', 'Product', 'DCFR Ranking']).size().unstack(fill_value=0)
    for r in dcfr_rankings:
        if r not in dcfr_prod_pivot.columns: dcfr_prod_pivot[r] = 0
    dcfr_prod_pivot = dcfr_prod_pivot[dcfr_rankings]
    dcfr_prod_pivot['Total'] = dcfr_prod_pivot.sum(axis=1)
    released_count = df_dcfr[(df_dcfr['Final Overall Status'] == 'Released (A)') & (df_dcfr['DCFR Ranking'].isin(['High','Medium','Larger Scale']))].groupby(['Pillar','Product']).size().reset_index(name='Released')
    hml_count = df_dcfr[df_dcfr['DCFR Ranking'].isin(['High','Medium','Larger Scale'])].groupby(['Pillar','Product']).size().reset_index(name='High+Med+Large Total')
    dcfr_prod_pivot = dcfr_prod_pivot.reset_index()
    dcfr_prod_pivot = dcfr_prod_pivot.merge(hml_count, on=['Pillar','Product'], how='left').merge(released_count, on=['Pillar','Product'], how='left')
    dcfr_prod_pivot['Released'] = dcfr_prod_pivot['Released'].fillna(0).astype(int)
    dcfr_prod_pivot['High+Med+Large Total'] = dcfr_prod_pivot['High+Med+Large Total'].fillna(0).astype(int)
    dcfr_prod_pivot['Release %'] = dcfr_prod_pivot.apply(lambda row: round(row['Released']/row['High+Med+Large Total']*100,1) if row['High+Med+Large Total']>0 else 0.0, axis=1)
    dcfr_prod_with_total = add_total_row(dcfr_prod_pivot, label_col='Pillar')
    def style_dcfr_prod(row):
        if row['Pillar'] == 'Total': return ['font-weight: bold; background-color: #f0f0f0'] * len(row)
        if isinstance(row['Release %'], (int, float)) and row['Release %'] < 50.0: return ['color: red; font-weight: bold'] * len(row)
        return [''] * len(row)
    styled_dcfr_prod = dcfr_prod_with_total.style.apply(style_dcfr_prod, axis=1).format({'Release %': lambda x: f'{x:.1f}%' if isinstance(x, (int, float)) else x}).hide(axis='index')
    st.dataframe(styled_dcfr_prod, use_container_width=True, column_config=col_config_compact(dcfr_prod_with_total))
    st.markdown('<p style="color:red; font-size:0.85rem; font-weight:bold;">⚠️ Release % is calculated only for High, Medium and Larger Scale features</p>', unsafe_allow_html=True)
    st.markdown("---")
    st.markdown('<div class="section-header">DCFR Feature Details</div>', unsafe_allow_html=True)
    reason_col = 'Reason why NF content cannot be built?'
    detail_cols = ['Pillar','Product','Module','Feature','CDL Name','DCFR Ranking','Training Required? ','Final Overall Status',reason_col]
    detail_cols_available = [c for c in detail_cols if c in df_dcfr.columns]
    detail = df_dcfr[detail_cols_available].copy().reset_index(drop=True)
    detail['Training Required? '] = detail['Training Required? '].fillna('Not Set')
    high_medium_large = ['High', 'Medium', 'Larger Scale']
    def style_dcfr_detail(row):
        training_val = row.get('Training Required? ', ''); reason_val = row.get(reason_col, ''); dcfr_val = row.get('DCFR Ranking', '')
        reason_blank = pd.isna(reason_val) or str(reason_val).strip() == ''
        if training_val == 'No' and reason_blank: return ['color: red; font-weight: bold'] * len(row)
        elif dcfr_val in high_medium_large and training_val in ['No', 'TBD', 'Not Set']: return ['color: darkorange; font-weight: bold'] * len(row)
        return [''] * len(row)
    if reason_col in detail.columns:
        st.dataframe(detail.style.apply(style_dcfr_detail, axis=1).hide(axis='index'), use_container_width=True, column_config=col_config_compact(detail))
    else:
        st.dataframe(detail, use_container_width=True, column_config=col_config_compact(detail))
    st.download_button("⬇️ Download DCFR Report", export_excel(detail), "report4_dcfr.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─────────────────────────────────────────────
# REPORT 5: CDL Coverage
# ─────────────────────────────────────────────
elif selected_report == "5️⃣  CDL Coverage — Pillar & CDL Wise":
    st.markdown('<div class="section-header">CDL Coverage — Pillar & CDL Wise</div>', unsafe_allow_html=True)
    df_filtered = render_report_filters(df_filtered, 'r5', has_cdl=True)
    df_r5 = df_filtered.copy()
    df_r5['CDL Name'] = df_r5['CDL Name'].fillna('Unassigned')
    df_r5['Training Required? '] = df_r5['Training Required? '].fillna('Not Set')
    df_r5_nf    = df_r5[df_r5['Feature Category'] != 'Unboxing'].copy()
    df_r5_unbox = df_r5[df_r5['Feature Category'] == 'Unboxing'].copy()
    def map_pillar_group_r5(pillar):
        if str(pillar).startswith('CX'): return 'CX'
        return pillar
    df_r5['Pillar Group']    = df_r5['Pillar'].apply(map_pillar_group_r5)
    df_r5_nf['Pillar Group'] = df_r5_nf['Pillar'].apply(map_pillar_group_r5)
    st.markdown('<div class="section-header">📦 Pillar Wise — CDL Summary</div>', unsafe_allow_html=True)
    pillar_groups = ['CX', 'ERP', 'HCM', 'PRC', 'SCM']
    box_cols = st.columns(5)
    for col, pg in zip(box_cols, pillar_groups):
        pg_nf    = df_r5_nf[df_r5_nf['Pillar Group'] == pg]
        pg_unbox = df_r5_unbox[df_r5_unbox['Pillar'].apply(map_pillar_group_r5) == pg]
        cdl_nf_counts    = pg_nf[pg_nf['Training Required? '] == 'Yes'].groupby('CDL Name').size()
        cdl_unbox_counts = pg_unbox.groupby('CDL Name').size()
        all_cdls = sorted(set(cdl_nf_counts.index.tolist() + cdl_unbox_counts.index.tolist()))
        all_cdls = [c for c in all_cdls if c != 'Unassigned']
        cdl_lines = []
        for cdl in all_cdls:
            nf_count    = cdl_nf_counts.get(cdl, 0)
            unbox_count = cdl_unbox_counts.get(cdl, 0)
            cdl_lines.append(f'<div style="font-size:0.78rem; color:#1a1a2e; margin-bottom:0.3rem;"><b>{cdl}</b>: <span style="color:#4f46e5">NF: {nf_count}</span> | <span style="color:#f59e0b">Unboxing: {unbox_count}</span></div>')
        cdl_content = ''.join(cdl_lines) if cdl_lines else '<div style="font-size:0.78rem; color:#94a3b8;">No CDLs assigned</div>'
        col.markdown(f'<div class="metric-card" style="border-left-color:#4f46e5; text-align:left; padding:1rem;"><div style="font-size:1rem; font-weight:700; color:#1a1a2e; margin-bottom:0.5rem;">{pg}</div>{cdl_content}</div>', unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown('<div class="section-header">Features per CDL — New Features vs Unboxing</div>', unsafe_allow_html=True)
    cdl_nf_all = df_r5_nf.groupby('CDL Name').size().reset_index(name='Count')
    cdl_nf_all['Type'] = 'New Features'
    cdl_unbox_all = df_r5_unbox.groupby('CDL Name').size().reset_index(name='Count')
    cdl_unbox_all['Type'] = 'Unboxing'
    cdl_combined = pd.concat([cdl_nf_all, cdl_unbox_all], ignore_index=True)
    cdl_combined = cdl_combined[cdl_combined['CDL Name'] != 'Unassigned']
    fig_cdl = px.bar(cdl_combined, x='CDL Name', y='Count', color='Type', barmode='group', title='CDL Assignment — New Features vs Unboxing', color_discrete_map={'New Features':'#4f46e5','Unboxing':'#f59e0b'}, text='Count')
    fig_cdl.update_traces(textposition='outside')
    fig_cdl.update_layout(plot_bgcolor='white', paper_bgcolor='white', xaxis_tickangle=-30, margin=dict(t=60, b=100), legend=dict(orientation='h', yanchor='top', y=-0.3, xanchor='center', x=0.5, title_text=''))
    st.plotly_chart(fig_cdl, use_container_width=True)
    cdl_nf_pivot    = df_r5_nf.groupby(['Pillar','CDL Name']).size().reset_index(name='New Features')
    cdl_unbox_pivot = df_r5_unbox.groupby(['Pillar','CDL Name']).size().reset_index(name='Unboxing')
    cdl_combined_table = cdl_nf_pivot.merge(cdl_unbox_pivot, on=['Pillar','CDL Name'], how='outer').fillna(0)
    cdl_combined_table['New Features'] = cdl_combined_table['New Features'].astype(int)
    cdl_combined_table['Unboxing']     = cdl_combined_table['Unboxing'].astype(int)
    cdl_combined_table['Total']        = cdl_combined_table['New Features'] + cdl_combined_table['Unboxing']
    cdl_combined_table = cdl_combined_table[cdl_combined_table['CDL Name'] != 'Unassigned']
    cdl_combined_with_total = add_total_row(cdl_combined_table, label_col='Pillar')
    st.dataframe(style_with_total(cdl_combined_with_total), use_container_width=True, column_config=col_config_compact(cdl_combined_with_total))
    st.markdown("---")
    st.markdown('<div class="section-header">📊 CDL Performance — New Features</div>', unsafe_allow_html=True)
    cdl_perf = df_r5_nf[df_r5_nf['CDL Name'] != 'Unassigned'].copy()
    cdl_total    = cdl_perf.groupby('CDL Name').size().reset_index(name='Total Assigned')
    cdl_yes      = cdl_perf[cdl_perf['Training Required? '] == 'Yes'].groupby('CDL Name').size().reset_index(name='Training = Yes')
    cdl_released = cdl_perf[cdl_perf['Final Overall Status'] == 'Released (A)'].groupby('CDL Name').size().reset_index(name='Released')
    cdl_perf_df  = cdl_total.merge(cdl_yes, on='CDL Name', how='left').merge(cdl_released, on='CDL Name', how='left').fillna(0)
    cdl_perf_df['Training = Yes'] = cdl_perf_df['Training = Yes'].astype(int)
    cdl_perf_df['Released']       = cdl_perf_df['Released'].astype(int)
    cdl_perf_melted = cdl_perf_df.melt(id_vars='CDL Name', value_vars=['Total Assigned','Training = Yes','Released'], var_name='Metric', value_name='Count')
    fig_perf = px.bar(cdl_perf_melted, x='CDL Name', y='Count', color='Metric', barmode='group', title='CDL Performance (Assigned Vs Training To Be Built) — New Features', color_discrete_map={'Total Assigned':'#4f46e5','Training = Yes':'#10b981','Released':'#f59e0b'}, text='Count')
    fig_perf.update_traces(textposition='outside')
    fig_perf.update_layout(plot_bgcolor='white', paper_bgcolor='white', xaxis_tickangle=-30, margin=dict(t=60, b=100), legend=dict(orientation='h', yanchor='top', y=-0.3, xanchor='center', x=0.5, title_text=''))
    st.plotly_chart(fig_perf, use_container_width=True)
    cdl_perf_df['Development %'] = cdl_perf_df.apply(lambda row: round(row['Training = Yes']/row['Total Assigned']*100,1) if row['Total Assigned']>0 else 0.0, axis=1)
    cdl_perf_df['Release %']     = cdl_perf_df.apply(lambda row: round(row['Released']/row['Training = Yes']*100,1) if row['Training = Yes']>0 else 0.0, axis=1)
    cdl_perf_with_total = add_total_row(cdl_perf_df, label_col='CDL Name')
    total_mask = cdl_perf_with_total['CDL Name'] == 'Total'
    t_assigned = cdl_perf_with_total.loc[total_mask, 'Total Assigned'].values[0]
    t_yes      = cdl_perf_with_total.loc[total_mask, 'Training = Yes'].values[0]
    t_rel      = cdl_perf_with_total.loc[total_mask, 'Released'].values[0]
    if t_assigned > 0: cdl_perf_with_total.loc[total_mask, 'Development %'] = round((t_yes/t_assigned)*100,1)
    if t_yes > 0: cdl_perf_with_total.loc[total_mask, 'Release %'] = round((t_rel/t_yes)*100,1)
    count_cols = ['Total Assigned', 'Training = Yes', 'Released']
    cdl_perf_with_total[count_cols] = cdl_perf_with_total[count_cols].astype(int)
    def style_cdl_perf(row):
        if row['CDL Name'] == 'Total': return ['font-weight: bold; background-color: #f0f0f0'] * len(row)
        styles = [''] * len(row); cols = list(row.index)
        if 'Development %' in cols and isinstance(row['Development %'], (int, float)) and row['Development %'] < 50.0:
            styles[cols.index('Development %')] = 'color: darkorange; font-weight: bold'
        if 'Release %' in cols and isinstance(row['Release %'], (int, float)) and row['Release %'] < 50.0:
            styles[cols.index('Release %')] = 'color: darkorange; font-weight: bold'
        return styles
    styled_perf = cdl_perf_with_total.style.apply(style_cdl_perf, axis=1).format({'Development %': lambda x: f'{x:.1f}%' if isinstance(x, (int, float)) else x, 'Release %': lambda x: f'{x:.1f}%' if isinstance(x, (int, float)) else x}).hide(axis='index')
    st.dataframe(styled_perf, use_container_width=True, column_config=col_config_compact(cdl_perf_with_total))
    st.markdown("---")
    st.markdown('<div class="section-header">📦 CDL Unboxing Coverage</div>', unsafe_allow_html=True)
    cdl_unbox_perf  = df_r5_unbox[df_r5_unbox['CDL Name'] != 'Unassigned'].copy()
    cdl_unbox_total = cdl_unbox_perf.groupby('CDL Name').size().reset_index(name='Total Unboxing')
    cdl_unbox_rel   = cdl_unbox_perf[cdl_unbox_perf['Final Overall Status'] == 'Released (A)'].groupby('CDL Name').size().reset_index(name='Released Unboxing')
    cdl_unbox_df    = cdl_unbox_total.merge(cdl_unbox_rel, on='CDL Name', how='left').fillna(0)
    cdl_unbox_df['Released Unboxing'] = cdl_unbox_df['Released Unboxing'].astype(int)
    cdl_unbox_df['Release %'] = cdl_unbox_df.apply(lambda row: round(row['Released Unboxing']/row['Total Unboxing']*100,1) if row['Total Unboxing']>0 else 0.0, axis=1)
    cdl_unbox_with_total = add_total_row(cdl_unbox_df, label_col='CDL Name')
    total_mask = cdl_unbox_with_total['CDL Name'] == 'Total'
    t_total_unbox = cdl_unbox_with_total.loc[total_mask, 'Total Unboxing'].values[0]
    t_rel_unbox   = cdl_unbox_with_total.loc[total_mask, 'Released Unboxing'].values[0]
    if t_total_unbox > 0: cdl_unbox_with_total.loc[total_mask, 'Release %'] = round((t_rel_unbox/t_total_unbox)*100,1)
    count_cols = ['Total Unboxing', 'Released Unboxing']
    cdl_unbox_with_total[count_cols] = cdl_unbox_with_total[count_cols].astype(int)
    def style_unbox_perf(row):
        if row['CDL Name'] == 'Total': return ['font-weight: bold; background-color: #f0f0f0'] * len(row)
        styles = [''] * len(row); cols = list(row.index)
        if 'Release %' in cols and isinstance(row['Release %'], (int, float)) and row['Release %'] < 50.0:
            styles[cols.index('Release %')] = 'color: darkorange; font-weight: bold'
        return styles
    styled_unbox = cdl_unbox_with_total.style.apply(style_unbox_perf, axis=1).format({'Release %': lambda x: f'{x:.1f}%' if isinstance(x, (int, float)) else x}).hide(axis='index')
    st.dataframe(styled_unbox, use_container_width=True, column_config=col_config_compact(cdl_unbox_with_total))
    st.markdown("---")
    st.markdown('<div class="section-header">CDL Feature Detail</div>', unsafe_allow_html=True)
    selected_cdl = st.selectbox("Filter by CDL", ['All'] + sorted(df_r5[df_r5['CDL Name'] != 'Unassigned']['CDL Name'].unique().tolist()), key='r5_cdl_detail')
    reason_col = 'Reason why NF content cannot be built?'
    detail_cols = ['Pillar','Product','Feature','CDL Name','Feature Category','Training Required? ','Final Overall Status',reason_col]
    detail_cols_available = [c for c in detail_cols if c in df_r5.columns]
    df_r5_detail = df_r5[detail_cols_available].copy()
    if selected_cdl != 'All': df_r5_detail = df_r5_detail[df_r5_detail['CDL Name'] == selected_cdl]
    def style_r5_detail(row):
        training_val = row.get('Training Required? ', ''); reason_val = row.get(reason_col, '')
        reason_blank = pd.isna(reason_val) or str(reason_val).strip() == ''
        if training_val == 'No' and reason_blank: return ['color: red; font-weight: bold'] * len(row)
        return [''] * len(row)
    if reason_col in df_r5_detail.columns:
        st.dataframe(df_r5_detail.style.apply(style_r5_detail, axis=1).hide(axis='index'), use_container_width=True, column_config=col_config_compact(df_r5_detail))
    else:
        st.dataframe(df_r5_detail.reset_index(drop=True), use_container_width=True, column_config=col_config_compact(df_r5_detail))
    st.download_button("⬇️ Download CDL Coverage Report", export_excel(df_r5_detail), "report5_cdl_coverage.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─────────────────────────────────────────────
# REPORT 6: CDL Recording Status
# ─────────────────────────────────────────────
elif selected_report == "6️⃣  CDL Recording Status — Pillar & CDL Wise":
    st.markdown('<div class="section-header">Recording Status — Pillar & CDL Wise</div>', unsafe_allow_html=True)
    df_filtered = render_report_filters(df_filtered, 'r6', has_cdl=True)
    import numpy as np
    today = pd.Timestamp(datetime.today().date())
    df_r6 = df_filtered.copy()
    df_r6['CDL Name'] = df_r6['CDL Name'].fillna('Unassigned')
    df_r6['Training Required? '] = df_r6['Training Required? '].fillna('Not Set')
    df_r6['Video Ready Date'] = pd.to_datetime(df_r6['Video Ready Date'], errors='coerce')
    df_r6['Recording Date']   = pd.to_datetime(df_r6['Recording Date'], errors='coerce')
    valid_from = pd.Timestamp('2020-01-01')
    df_r6['Video Ready Date'] = df_r6['Video Ready Date'].apply(lambda x: x if pd.notna(x) and x >= valid_from else pd.NaT)
    df_r6['Recording Date']   = df_r6['Recording Date'].apply(lambda x: x if pd.notna(x) and x >= valid_from else pd.NaT)
    fc_col = 'Feature Classification'
    completed_primary_map = df_r6.groupby('CDL Name').apply(
        lambda g: set(g[(g[fc_col] == 'Combined Primary NF') & (g['Video Ready Date'].notna()) & (g['Combined Feature Video Name'].notna())]['Combined Feature Video Name'])
    ).to_dict()
    def get_recording_status(row):
        fc   = str(row.get(fc_col, '') or '').strip()
        vrd  = row.get('Video Ready Date', None)
        cfvn = str(row.get('Combined Feature Video Name', '') or '').strip()
        cdl  = str(row.get('CDL Name', '') or '').strip()
        if fc == 'Combined Bundled NF':
            return 'Complete' if cfvn in completed_primary_map.get(cdl, set()) else 'Pending'
        return 'Complete' if pd.notna(vrd) else 'Pending'
    df_r6['Recording Status'] = df_r6.apply(get_recording_status, axis=1)
    df_r6_nf    = df_r6[df_r6['Feature Category'] != 'Unboxing'].copy()
    df_r6_unbox = df_r6[df_r6['Feature Category'] == 'Unboxing'].copy()
    def map_pillar_group_r6(pillar):
        if str(pillar).startswith('CX'): return 'CX'
        return pillar
    df_r6['Pillar Group'] = df_r6['Pillar'].apply(map_pillar_group_r6)
    st.markdown('<div class="section-header">📦 Pillar Wise — Recording Status</div>', unsafe_allow_html=True)
    pillar_groups = ['CX', 'ERP', 'HCM', 'PRC', 'SCM']
    box_cols = st.columns(5)
    for col, pg in zip(box_cols, pillar_groups):
        pg_df = df_r6[df_r6['Pillar Group'] == pg]
        pg_nf = df_r6_nf[df_r6_nf['Pillar'].apply(map_pillar_group_r6) == pg]
        pg_ub = df_r6_unbox[df_r6_unbox['Pillar'].apply(map_pillar_group_r6) == pg]
        total      = len(pg_df)
        pg_yes     = pg_df[pg_df['Training Required? '] == 'Yes']
        yes_total  = len(pg_yes)
        complete   = len(pg_yes[pg_yes['Recording Status'] == 'Complete'])
        pending    = yes_total - complete
        comp_pct   = round(complete / yes_total * 100, 1) if yes_total > 0 else 0.0
        nf_yes_count = len(pg_nf[pg_nf['Training Required? '] == 'Yes'])
        nf_released  = len(pg_nf[(pg_nf['Training Required? '] == 'Yes') & (pg_nf['Final Overall Status'] == 'Released (A)')])
        nf_rel_pct   = round(nf_released / nf_yes_count * 100, 1) if nf_yes_count > 0 else 0.0
        nf_rel_color = '#10b981' if nf_rel_pct >= 50 else '#ef4444'
        ub_yes_count = len(pg_ub[pg_ub['Training Required? '] == 'Yes'])
        ub_released  = len(pg_ub[(pg_ub['Training Required? '] == 'Yes') & (pg_ub['Final Overall Status'] == 'Released (A)')])
        ub_rel_pct   = round(ub_released / ub_yes_count * 100, 1) if ub_yes_count > 0 else 0.0
        ub_rel_color = '#10b981' if ub_rel_pct >= 50 else '#ef4444'
        col.markdown(
            f'<div class="metric-card" style="border-left-color:#4f46e5; text-align:left; padding:1rem;">'
            f'<div style="font-size:1rem; font-weight:700; color:#1a1a2e; margin-bottom:0.5rem;">{pg}</div>'
            f'<div style="font-size:0.85rem; color:#6c757d;">Total Features: <b>{total}</b></div>'
            f'<div style="font-size:0.85rem; color:#4f46e5; margin-bottom:0.3rem; font-weight:600;">Training = Yes: <b>{yes_total}</b></div>'
            f'<div style="font-size:0.85rem; color:#10b981;">Complete: <b>{complete}</b></div>'
            f'<div style="font-size:0.85rem; color:#ef4444;">Pending: <b>{pending}</b></div>'
            f'<div style="font-size:0.85rem; color:#4f46e5;">Completion: <b>{comp_pct}%</b></div>'
            f'<div style="border-top:1px solid #e9ecef; margin:0.4rem 0;"></div>'
            f'<div style="font-size:0.82rem; font-weight:700; color:{nf_rel_color};">🚀 NF Release %: {nf_rel_pct}%</div>'
            f'<div style="font-size:0.75rem; color:#94a3b8; margin-bottom:0.2rem;">({nf_released} of {nf_yes_count} Training=Yes NF)</div>'
            f'<div style="font-size:0.82rem; font-weight:700; color:{ub_rel_color};">📦 Unboxing Release %: {ub_rel_pct}%</div>'
            f'<div style="font-size:0.75rem; color:#94a3b8;">({ub_released} of {ub_yes_count} Training=Yes Unboxing)</div>'
            f'</div>', unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown('<div class="section-header">Recording Status by Pillar</div>', unsafe_allow_html=True)
    df_r6_nf_yes    = df_r6_nf[df_r6_nf['Training Required? '] == 'Yes'].copy()
    df_r6_unbox_yes = df_r6_unbox[df_r6_unbox['Training Required? '] == 'Yes'].copy()
    pillar_nf_rec = df_r6_nf_yes.groupby(['Pillar','Recording Status']).size().reset_index(name='Count')
    pillar_nf_rec['Type'] = 'New Features'
    pillar_ub_rec = df_r6_unbox_yes.groupby(['Pillar','Recording Status']).size().reset_index(name='Count')
    pillar_ub_rec['Type'] = 'Unboxing'
    pillar_rec_combined = pd.concat([pillar_nf_rec, pillar_ub_rec], ignore_index=True)
    pillar_rec_combined['Legend'] = pillar_rec_combined['Type'] + ' — ' + pillar_rec_combined['Recording Status']
    COLOR_MAP_R6 = {'New Features — Complete':'#10b981','New Features — Pending':'#ef4444','Unboxing — Complete':'#4f46e5','Unboxing — Pending':'#f59e0b'}
    fig_pillar = px.bar(pillar_rec_combined, x='Pillar', y='Count', color='Legend', barmode='group',
                        title='Recording Status by Pillar — New Features & Unboxing (Training = Yes only)',
                        color_discrete_map=COLOR_MAP_R6, text='Count')
    fig_pillar.update_traces(textposition='outside')
    fig_pillar.update_layout(plot_bgcolor='white', paper_bgcolor='white', margin=dict(t=60, b=80),
                             legend=dict(orientation='h', yanchor='top', y=-0.2, xanchor='center', x=0.5, title_text=''))
    st.plotly_chart(fig_pillar, use_container_width=True)
    st.markdown("---")
    st.markdown('<div class="section-header">CDL Recording Status — New Features</div>', unsafe_allow_html=True)
    cdl_nf = df_r6_nf[df_r6_nf['CDL Name'] != 'Unassigned'].copy()
    cdl_nf_yes      = cdl_nf[cdl_nf['Training Required? '] == 'Yes'].groupby('CDL Name').size().reset_index(name='Training = Yes')
    cdl_nf_complete = cdl_nf[(cdl_nf['Training Required? '] == 'Yes') & (cdl_nf['Recording Status'] == 'Complete')].groupby('CDL Name').size().reset_index(name='Recordings Complete')
    cdl_nf_chart    = cdl_nf_yes.merge(cdl_nf_complete, on='CDL Name', how='left').fillna(0)
    cdl_nf_chart['Recordings Complete'] = cdl_nf_chart['Recordings Complete'].astype(int)
    cdl_nf_melted = cdl_nf_chart.melt(id_vars='CDL Name', value_vars=['Training = Yes','Recordings Complete'], var_name='Metric', value_name='Count')
    fig_cdl_nf = px.bar(cdl_nf_melted, x='CDL Name', y='Count', color='Metric', barmode='group', title='CDL Recording Status — New Features', color_discrete_map={'Training = Yes':'#4f46e5','Recordings Complete':'#10b981'}, text='Count')
    fig_cdl_nf.update_traces(textposition='outside')
    fig_cdl_nf.update_layout(plot_bgcolor='white', paper_bgcolor='white', xaxis_tickangle=-30, margin=dict(t=60, b=100), legend=dict(orientation='h', yanchor='top', y=-0.3, xanchor='center', x=0.5, title_text=''))
    st.plotly_chart(fig_cdl_nf, use_container_width=True)
    st.markdown("---")
    st.markdown('<div class="section-header">CDL Recording Status — Unboxing</div>', unsafe_allow_html=True)
    cdl_unbox = df_r6_unbox[df_r6_unbox['CDL Name'] != 'Unassigned'].copy()
    cdl_unbox_total    = cdl_unbox.groupby('CDL Name').size().reset_index(name='Total Unboxing')
    cdl_unbox_complete = cdl_unbox[cdl_unbox['Recording Status'] == 'Complete'].groupby('CDL Name').size().reset_index(name='Recordings Complete')
    cdl_unbox_chart    = cdl_unbox_total.merge(cdl_unbox_complete, on='CDL Name', how='left').fillna(0)
    cdl_unbox_chart['Recordings Complete'] = cdl_unbox_chart['Recordings Complete'].astype(int)
    cdl_unbox_chart = cdl_unbox_chart[cdl_unbox_chart['Total Unboxing'] > 0]
    if len(cdl_unbox_chart) == 0:
        st.info("No Unboxing recordings found.")
    else:
        cdl_unbox_melted = cdl_unbox_chart.melt(id_vars='CDL Name', value_vars=['Total Unboxing','Recordings Complete'], var_name='Metric', value_name='Count')
        fig_cdl_unbox = px.bar(cdl_unbox_melted, x='CDL Name', y='Count', color='Metric', barmode='group', title='CDL Recording Status — Unboxing', color_discrete_map={'Total Unboxing':'#7c3aed','Recordings Complete':'#10b981'}, text='Count')
        fig_cdl_unbox.update_traces(textposition='outside')
        fig_cdl_unbox.update_layout(plot_bgcolor='white', paper_bgcolor='white', xaxis_tickangle=-30, margin=dict(t=60, b=100), legend=dict(orientation='h', yanchor='top', y=-0.3, xanchor='center', x=0.5, title_text=''))
        st.plotly_chart(fig_cdl_unbox, use_container_width=True)
    st.markdown("---")
    st.markdown('<div class="section-header">📋 CDL Recording Timeline & Criteria Summary</div>', unsafe_allow_html=True)
    cdl_timeline     = df_r6_nf[df_r6_nf['CDL Name'] != 'Unassigned'].copy()
    cdl_pillar       = cdl_timeline.groupby('CDL Name')['Pillar'].first().reset_index()
    cdl_dates        = cdl_timeline.groupby('CDL Name')['Recording Date'].agg(Start_Recording_Date='min', End_Recording_Date='max').reset_index()
    cdl_nf_yes_count = cdl_timeline[cdl_timeline['Training Required? '] == 'Yes'].groupby('CDL Name').size().reset_index(name='Total NF (Yes)')
    cdl_no           = cdl_timeline[cdl_timeline['Training Required? '] == 'No'].groupby('CDL Name').size().reset_index(name='Training = No')
    cdl_tbd          = cdl_timeline[cdl_timeline['Training Required? '] == 'TBD'].groupby('CDL Name').size().reset_index(name='Training = TBD')
    cdl_blank        = cdl_timeline[cdl_timeline['Training Required? '] == 'Not Set'].groupby('CDL Name').size().reset_index(name='Training = Blank')
    if 'CDL Daily Status Update' in cdl_timeline.columns:
        cdl_status = cdl_timeline.groupby('CDL Name')['CDL Daily Status Update'].last().reset_index()
    else:
        cdl_status = cdl_timeline.groupby('CDL Name')['CDL Name'].first().reset_index()
        cdl_status['CDL Daily Status Update'] = 'Not Available'
    summary = (cdl_pillar.merge(cdl_dates, on='CDL Name', how='left').merge(cdl_nf_yes_count, on='CDL Name', how='left')
               .merge(cdl_no, on='CDL Name', how='left').merge(cdl_tbd, on='CDL Name', how='left')
               .merge(cdl_blank, on='CDL Name', how='left').merge(cdl_status, on='CDL Name', how='left').fillna(0))
    summary['Total NF (Yes)']   = summary['Total NF (Yes)'].astype(int)
    summary['Training = No']    = summary['Training = No'].astype(int)
    summary['Training = TBD']   = summary['Training = TBD'].astype(int)
    summary['Training = Blank'] = summary['Training = Blank'].astype(int)
    summary['Start_Recording_Date'] = pd.to_datetime(summary['Start_Recording_Date'], errors='coerce')
    summary['End_Recording_Date']   = pd.to_datetime(summary['End_Recording_Date'], errors='coerce')
    summary['Start_Recording_Date'] = summary['Start_Recording_Date'].apply(lambda x: x if pd.notna(x) and x >= valid_from else pd.NaT)
    summary['End_Recording_Date']   = summary['End_Recording_Date'].apply(lambda x: x if pd.notna(x) and x >= valid_from else pd.NaT)
    def calc_working_days(row):
        if pd.isna(row['Start_Recording_Date']) or pd.isna(row['End_Recording_Date']): return 'Not Set'
        days = np.busday_count(row['Start_Recording_Date'].date(), row['End_Recording_Date'].date())
        return max(int(days), 1)
    def calc_features_per_day(row):
        days = row['Total Days Blocked']; nf = row['Total NF (Yes)']
        if days == 'Not Set': return 'Not Set'
        elif days == 0: return str(round(float(nf), 1))
        else: return str(round(nf / days, 1))
    summary['Total Days Blocked'] = summary.apply(calc_working_days, axis=1)
    summary['No Of Features CDL is scheduled to develop in a day'] = summary.apply(calc_features_per_day, axis=1)
    summary['Start Recording Date'] = summary['Start_Recording_Date'].apply(lambda x: x.strftime('%d-%b-%Y') if pd.notna(x) else 'Not Set')
    summary['End Recording Date']   = summary['End_Recording_Date'].apply(lambda x: x.strftime('%d-%b-%Y') if pd.notna(x) else 'Not Set')
    final_cols = ['Pillar','CDL Name','Start Recording Date','End Recording Date','Total NF (Yes)','Total Days Blocked','No Of Features CDL is scheduled to develop in a day','Training = No','Training = TBD','Training = Blank','CDL Daily Status Update']
    summary_display = summary[[c for c in final_cols if c in summary.columns]]
    summary_display = summary_display[summary_display['Total NF (Yes)'] != 0].copy()
    def style_timeline(row):
        start = row.get('Start Recording Date', 'Not Set'); end = row.get('End Recording Date', 'Not Set'); nf = row.get('Total NF (Yes)', 0)
        if start != 'Not Set' and end != 'Not Set' and start == end and nf > 1:
            return ['background-color: #fff3cd; color: #856404; font-weight: bold'] * len(row)
        return [''] * len(row)
    styled_timeline = summary_display.style.apply(style_timeline, axis=1).hide(axis='index')
    st.dataframe(styled_timeline, use_container_width=True, column_config=col_config_compact(summary_display))
    st.markdown('<p style="color:#856404; font-size:0.85rem;">⚠️ <b>Orange rows</b> indicate CDLs where Start and End Recording Date are the same but have more than 1 feature.</p>', unsafe_allow_html=True)
    st.download_button("⬇️ Download Recording Timeline Report", export_excel(summary_display), "report6_recording_timeline.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    st.markdown("---")
    st.markdown('<div class="section-header">Recording Status Detail</div>', unsafe_allow_html=True)
    _GA_BASE_R6 = pd.Timestamp('2026-07-03')
    def get_ga_tier_r6(rec_date):
        if pd.isna(rec_date): return 'Not Set'
        release_date = rec_date + pd.Timedelta(days=4)
        if release_date <= _GA_BASE_R6: return 'GA+0'
        diff = (release_date - _GA_BASE_R6).days
        tier = ((diff - 1) // 7) + 1
        return f'GA+{tier}'
    def calc_recording_status_detail(row):
        rec_date = row['Recording Date']; rs = row['Recording Status']
        if pd.isna(rec_date): return 'Yet to receive recording dates'
        elif rs == 'Complete': return 'Complete'
        elif pd.notna(rec_date) and today > rec_date: return 'Delayed'
        else: return 'Upcoming'
    df_r6['Recording Status Detail'] = df_r6.apply(calc_recording_status_detail, axis=1)
    df_r6['Feature Type'] = df_r6['Feature Category'].apply(lambda x: 'Unboxing' if str(x).strip() == 'Unboxing' else 'New Feature')
    detail_cols = ['Pillar', 'Product', 'Feature', 'CDL Name', 'Feature Type', 'Tags ( AI / Redwood/AI Agent)', 'Recording Date', 'Video Ready Date', 'Recording Status Detail', 'Final Overall Status', 'CDL Daily Status Update', 'Type of Issue']
    detail_cols_available = [c for c in detail_cols if c in df_r6.columns]
    detail = df_r6[df_r6['Training Required? '].isin(['Yes', 'TBD'])][detail_cols_available].copy()
    detail = detail.rename(columns={'Recording Status Detail': 'Recording Status', 'Tags ( AI / Redwood/AI Agent)': 'Feature Tag'})
    detail['Expected GA+ Release'] = detail['Recording Date'].apply(get_ga_tier_r6)
    cols_order = ['Pillar', 'Product', 'Feature', 'CDL Name', 'Feature Type', 'Feature Tag', 'Recording Date', 'Expected GA+ Release', 'Video Ready Date', 'Recording Status', 'Final Overall Status', 'CDL Daily Status Update', 'Type of Issue']
    detail = detail[[c for c in cols_order if c in detail.columns]]
    filter_status = st.selectbox("Filter by Status", ['All', 'Complete', 'Delayed', 'Upcoming', 'Yet to receive recording dates'], key='filter_rec_status_detail')
    if filter_status != 'All': detail = detail[detail['Recording Status'] == filter_status]
    def style_rec_detail(row):
        status = row.get('Recording Status', '')
        if status == 'Delayed': return ['color: red; font-weight: bold'] * len(row)
        elif status == 'Complete': return ['color: #10b981; font-weight: bold'] * len(row)
        elif status == 'Yet to receive recording dates': return ['color: #f59e0b; font-weight: bold'] * len(row)
        elif status == 'Upcoming': return ['color: #4f46e5; font-weight: bold'] * len(row)
        return [''] * len(row)
    styled_detail = detail.style.apply(style_rec_detail, axis=1).hide(axis='index')
    st.dataframe(styled_detail, use_container_width=True, column_config=col_config_compact(detail))
    st.markdown('<p style="font-size:0.85rem;"><span style="color:#10b981; font-weight:bold;">● Complete</span> — Video Ready Date submitted &nbsp;|&nbsp; <span style="color:red; font-weight:bold;">● Delayed</span> — Past Recording Date, no video submitted &nbsp;|&nbsp; <span style="color:#4f46e5; font-weight:bold;">● Upcoming</span> — Recording Date set, not yet due &nbsp;|&nbsp; <span style="color:#f59e0b; font-weight:bold;">● Yet to receive recording dates</span> — No Recording Date set</p>', unsafe_allow_html=True)
    st.markdown('<div style="background-color:#e8f4f8; padding:8px 14px; border-radius:6px; border-left:4px solid #0078d4; margin-top:6px; font-size:0.82rem; color:#444;">📌 <b>Expected GA+ Release</b> = Recording Date + 4 days, calculated against GA+0 base date <b>03-Jul-2026</b>. Use this to identify AI / Redwood tagged features and their expected release tier.</div>', unsafe_allow_html=True)
    st.markdown('<div style="background-color:#fffbeb; padding:8px 14px; border-radius:6px; border-left:4px solid #f59e0b; margin-top:6px; font-size:0.82rem; color:#444;">📌 <b>Note:</b> This table includes features where <b>Training Required = Yes</b> and <b>Training Required = TBD</b>. All other reports filter to Training = Yes only.</div>', unsafe_allow_html=True)
    st.download_button("⬇️ Download Recording Status Report", export_excel(detail), "report6_recording_status.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─────────────────────────────────────────────
# REPORT 7: Unboxing Videos
# ─────────────────────────────────────────────
elif selected_report == "7️⃣  Unboxing Videos — Pillar & Product Wise":
    st.markdown('<div class="section-header">Unboxing Videos — Pillar & Product Wise</div>', unsafe_allow_html=True)
    df_filtered = render_report_filters(df_filtered, 'r7', has_cdl=True)
    df_r7_unbox = df_filtered[df_filtered['Feature Category'] == 'Unboxing'].copy()
    df_r7_nf    = df_filtered[df_filtered['Feature Category'] != 'Unboxing'].copy()
    def map_pillar_group_r7(pillar):
        if str(pillar).startswith('CX'): return 'CX'
        return pillar
    df_r7_unbox['Pillar Group'] = df_r7_unbox['Pillar'].apply(map_pillar_group_r7)
    df_r7_nf['Pillar Group']    = df_r7_nf['Pillar'].apply(map_pillar_group_r7)
    st.markdown('<div class="section-header">📦 Pillar Wise — Unboxing Summary</div>', unsafe_allow_html=True)
    pillar_groups = ['CX', 'ERP', 'HCM', 'PRC', 'SCM']
    box_cols = st.columns(5)
    for col, pg in zip(box_cols, pillar_groups):
        pg_unbox     = df_r7_unbox[df_r7_unbox['Pillar Group'] == pg]
        total        = len(pg_unbox)
        released     = len(pg_unbox[pg_unbox['Final Overall Status'] == 'Released (A)'])
        pending      = total - released
        # ── CHANGE 1B: Training=Yes count shown, Release % uses Training=Yes ──
        ub_yes_count = len(pg_unbox[pg_unbox['Training Required? '] == 'Yes'])
        ub_rel_yes   = len(pg_unbox[(pg_unbox['Training Required? '] == 'Yes') & (pg_unbox['Final Overall Status'] == 'Released (A)')])
        ub_rel_pct   = round(ub_rel_yes / ub_yes_count * 100, 1) if ub_yes_count > 0 else 0.0
        ub_rel_color = '#10b981' if ub_rel_pct >= 50 else '#ef4444'
        col.markdown(
            f'<div class="metric-card" style="border-left-color:#7c3aed; text-align:left; padding:1rem;">'
            f'<div style="font-size:1rem; font-weight:700; color:#1a1a2e; margin-bottom:0.5rem;">{pg}</div>'
            f'<div style="font-size:0.85rem; color:#6c757d;">Total: <b>{total}</b></div>'
            f'<div style="font-size:0.85rem; color:#4f46e5;">Training = Yes: <b>{ub_yes_count}</b></div>'
            f'<div style="font-size:0.85rem; color:#10b981;">Released: <b>{released}</b></div>'
            f'<div style="font-size:0.85rem; color:#ef4444;">Pending: <b>{pending}</b></div>'
            f'<div style="font-size:0.82rem; font-weight:700; color:{ub_rel_color}; margin-top:0.3rem;">🚀 Release %: {ub_rel_pct}%</div>'
            f'<div style="font-size:0.75rem; color:#94a3b8;">({ub_rel_yes} of {ub_yes_count} Training=Yes)</div>'
            f'</div>', unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    # ── FIX (missing-unboxing warning): the old merge compared Product strings
    #    byte-for-byte, so a trailing space / double space / non-breaking space /
    #    casing difference between NF rows and Unboxing rows made a product look
    #    "missing" even though it appears in the detail table below.
    #    We now match on a normalized key: strip, collapse internal whitespace,
    #    replace non-breaking spaces, and compare case-insensitively. ──
    def _norm_product_r7(x):
        if pd.isna(x): return ''
        return ' '.join(str(x).replace('\u00a0', ' ').split()).strip().lower()
    all_products   = df_r7_nf[['Pillar', 'Product']].drop_duplicates().copy()
    unbox_products = df_r7_unbox[['Pillar', 'Product']].drop_duplicates().copy()
    all_products['_pkey']   = all_products['Product'].apply(_norm_product_r7)
    unbox_products['_pkey'] = unbox_products['Product'].apply(_norm_product_r7)
    missing_unbox = all_products.merge(
        unbox_products[['Pillar', '_pkey']].drop_duplicates(),
        on=['Pillar', '_pkey'], how='left', indicator=True
    )
    missing_unbox = missing_unbox[missing_unbox['_merge'] == 'left_only'][['Pillar', 'Product']]
    if len(missing_unbox) > 0:
        training_yes_count = df_r7_nf[df_r7_nf['Training Required? '] == 'Yes'].groupby(['Pillar', 'Product']).size().reset_index(name='Features (Training = Yes)')
        missing_unbox_table = missing_unbox.merge(training_yes_count, on=['Pillar', 'Product'], how='left')
        missing_unbox_table['Features (Training = Yes)'] = missing_unbox_table['Features (Training = Yes)'].fillna(0).astype(int)
        missing_unbox_table = missing_unbox_table[missing_unbox_table['Features (Training = Yes)'] > 0]
        if len(missing_unbox_table) > 0:
            st.markdown('<div class="upload-note">⚠️ <b>No Unboxing video found for the following products that have features with Training = Yes:</b></div>', unsafe_allow_html=True)
            st.dataframe(missing_unbox_table.reset_index(drop=True), use_container_width=True)
    st.markdown("---")
    st.markdown('<div class="section-header">Unboxing Videos by Pillar</div>', unsafe_allow_html=True)
    pillar_unbox_total    = df_r7_unbox.groupby('Pillar').size().reset_index(name='Count')
    pillar_unbox_total['Metric'] = 'Total Unboxing'
    pillar_unbox_released = df_r7_unbox[df_r7_unbox['Final Overall Status'] == 'Released (A)'].groupby('Pillar').size().reset_index(name='Count')
    pillar_unbox_released['Metric'] = 'Released Unboxing'
    pillar_unbox_combined = pd.concat([pillar_unbox_total, pillar_unbox_released], ignore_index=True)
    fig_pillar = px.bar(pillar_unbox_combined, x='Pillar', y='Count', color='Metric', barmode='group', title='Unboxing Videos by Pillar — Total vs Released', color_discrete_map={'Total Unboxing':'#7c3aed','Released Unboxing':'#10b981'}, text='Count')
    fig_pillar.update_traces(textposition='outside')
    fig_pillar.update_layout(plot_bgcolor='white', paper_bgcolor='white', margin=dict(t=60, b=80), legend=dict(orientation='h', yanchor='top', y=-0.2, xanchor='center', x=0.5, title_text=''))
    st.plotly_chart(fig_pillar, use_container_width=True)
    st.markdown("---")
    st.markdown('<div class="section-header">Unboxing Videos by Product</div>', unsafe_allow_html=True)
    prod_unbox_total    = df_r7_unbox.groupby(['Pillar','Product']).size().reset_index(name='Count')
    prod_unbox_total['Metric'] = 'Total Unboxing'
    prod_unbox_released = df_r7_unbox[df_r7_unbox['Final Overall Status'] == 'Released (A)'].groupby(['Pillar','Product']).size().reset_index(name='Count')
    prod_unbox_released['Metric'] = 'Released Unboxing'
    prod_unbox_combined = pd.concat([prod_unbox_total, prod_unbox_released], ignore_index=True)
    fig_prod = px.bar(prod_unbox_combined, x='Product', y='Count', color='Metric', barmode='group', title='Unboxing Videos by Product — Total vs Released', color_discrete_map={'Total Unboxing':'#7c3aed','Released Unboxing':'#10b981'}, text='Count')
    fig_prod.update_traces(textposition='outside')
    fig_prod.update_layout(plot_bgcolor='white', paper_bgcolor='white', xaxis_tickangle=-35, margin=dict(t=60, b=120), legend=dict(orientation='h', yanchor='top', y=-0.35, xanchor='center', x=0.5, title_text=''))
    st.plotly_chart(fig_prod, use_container_width=True)
    st.markdown("---")
    st.markdown('<div class="section-header">Unboxing Video Details</div>', unsafe_allow_html=True)
    detail_cols = ['Pillar','Product','Module','Feature','CDL Name','Final Overall Status','Video Ready Date']
    detail_cols_available = [c for c in detail_cols if c in df_r7_unbox.columns]
    detail = df_r7_unbox[detail_cols_available].reset_index(drop=True)
    st.dataframe(detail, use_container_width=True, column_config=col_config_compact(detail))
    st.download_button("⬇️ Download Unboxing Report", export_excel(detail), "report7_unboxing.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
# ─────────────────────────────────────────────
# REPORT 8: Released Videos
# ─────────────────────────────────────────────
elif selected_report == "8️⃣  Released Videos (incl. Unboxing)":
    st.markdown('<div class="section-header">Released Videos — Pillar Wise (Unboxing Separate)</div>', unsafe_allow_html=True)
    df_filtered = render_report_filters(df_filtered, 'r8', has_cdl=True)
    df_r8_nf    = df_filtered[df_filtered['Feature Category'] != 'Unboxing'].copy()
    df_r8_unbox = df_filtered[df_filtered['Feature Category'] == 'Unboxing'].copy()
    df_released  = df_filtered[df_filtered['Final Overall Status'] == 'Released (A)'].copy()
    df_rel_nf    = df_released[df_released['Feature Category'] != 'Unboxing']
    df_rel_unbox = df_released[df_released['Feature Category'] == 'Unboxing']
    def map_pillar_group_r8(pillar):
        if str(pillar).startswith('CX'): return 'CX'
        return pillar
    df_r8_nf['Pillar Group']    = df_r8_nf['Pillar'].apply(map_pillar_group_r8)
    df_r8_unbox['Pillar Group'] = df_r8_unbox['Pillar'].apply(map_pillar_group_r8)
    col1, col2, col3 = st.columns(3)
    col1.markdown(metric_card(len(df_released), "Total Released"), unsafe_allow_html=True)
    col2.markdown(metric_card(len(df_rel_nf), "Released (New Features)"), unsafe_allow_html=True)
    col3.markdown(metric_card(len(df_rel_unbox), "Released (Unboxing)"), unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="section-header">📦 Pillar Wise — Released Summary</div>', unsafe_allow_html=True)
    pillar_groups = ['CX', 'ERP', 'HCM', 'PRC', 'SCM']
    box_cols = st.columns(5)
    for col, pg in zip(box_cols, pillar_groups):
        pg_nf    = df_r8_nf[df_r8_nf['Pillar Group'] == pg]
        pg_unbox = df_r8_unbox[df_r8_unbox['Pillar Group'] == pg]
        nf_yes         = len(pg_nf[pg_nf['Training Required? '] == 'Yes'])
        nf_released    = len(pg_nf[pg_nf['Final Overall Status'] == 'Released (A)'])
        nf_rel_pct     = round(nf_released / nf_yes * 100, 1) if nf_yes > 0 else 0.0
        # ── Unboxing Release % uses Training=Yes ──
        ub_yes_count   = len(pg_unbox[pg_unbox['Training Required? '] == 'Yes'])
        unbox_released = len(pg_unbox[pg_unbox['Final Overall Status'] == 'Released (A)'])
        ub_rel_yes     = len(pg_unbox[(pg_unbox['Training Required? '] == 'Yes') & (pg_unbox['Final Overall Status'] == 'Released (A)')])
        ub_rel_pct     = round(ub_rel_yes / ub_yes_count * 100, 1) if ub_yes_count > 0 else 0.0
        nf_rel_color   = '#10b981' if nf_rel_pct >= 50 else '#ef4444'
        ub_rel_color   = '#10b981' if ub_rel_pct >= 50 else '#ef4444'
        col.markdown(
            f'<div class="metric-card" style="border-left-color:#4f46e5; text-align:left; padding:1rem;">'
            f'<div style="font-size:1rem; font-weight:700; color:#1a1a2e; margin-bottom:0.5rem;">{pg}</div>'
            f'<div style="font-size:0.78rem; color:#4f46e5;">NF Released: <b>{nf_released}</b></div>'
            f'<div style="font-size:0.78rem; font-weight:700; color:{nf_rel_color};">🚀 NF Release %: {nf_rel_pct}%</div>'
            f'<div style="font-size:0.72rem; color:#94a3b8; margin-bottom:0.3rem;">({nf_released} of {nf_yes} Training=Yes NF)</div>'
            f'<div style="font-size:0.78rem; color:#7c3aed;">Unboxing Released: <b>{unbox_released}</b></div>'
            f'<div style="font-size:0.78rem; font-weight:700; color:{ub_rel_color};">📦 Unboxing Release %: {ub_rel_pct}%</div>'
            f'<div style="font-size:0.72rem; color:#94a3b8;">({ub_rel_yes} of {ub_yes_count} Training=Yes Unboxing)</div>'
            f'</div>', unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown('<div class="section-header">🆕 New Features Released — Pillar Wise</div>', unsafe_allow_html=True)
    if len(df_rel_nf) == 0:
        st.info("No New Features released yet.")
    else:
        nf_yes_pillar = df_r8_nf[df_r8_nf['Training Required? '] == 'Yes'].groupby('Pillar').size().reset_index(name='Count')
        nf_yes_pillar['Metric'] = 'Training = Yes'
        nf_rel_pillar = df_rel_nf.groupby('Pillar').size().reset_index(name='Count')
        nf_rel_pillar['Metric'] = 'Released'
        nf_pillar_combined = pd.concat([nf_yes_pillar, nf_rel_pillar], ignore_index=True)
        fig_nf = px.bar(nf_pillar_combined, x='Pillar', y='Count', color='Metric', barmode='group', title='New Features — Training=Yes vs Released by Pillar', color_discrete_map={'Training = Yes':'#4f46e5','Released':'#10b981'}, text='Count')
        fig_nf.update_traces(textposition='outside')
        fig_nf.update_layout(plot_bgcolor='white', paper_bgcolor='white', margin=dict(t=60, b=80), legend=dict(orientation='h', yanchor='top', y=-0.2, xanchor='center', x=0.5, title_text=''))
        st.plotly_chart(fig_nf, use_container_width=True)
    st.markdown("---")
    st.markdown('<div class="section-header">📦 Unboxing Videos Released — Pillar Wise</div>', unsafe_allow_html=True)
    if len(df_rel_unbox) == 0:
        st.info("No Unboxing videos released yet.")
    else:
        unbox_total_pillar = df_r8_unbox.groupby('Pillar').size().reset_index(name='Count')
        unbox_total_pillar['Metric'] = 'Total Unboxing'
        unbox_rel_pillar   = df_rel_unbox.groupby('Pillar').size().reset_index(name='Count')
        unbox_rel_pillar['Metric'] = 'Released'
        unbox_pillar_combined = pd.concat([unbox_total_pillar, unbox_rel_pillar], ignore_index=True)
        fig_unbox = px.bar(unbox_pillar_combined, x='Pillar', y='Count', color='Metric', barmode='group', title='Unboxing Videos — Total vs Released by Pillar', color_discrete_map={'Total Unboxing':'#7c3aed','Released':'#10b981'}, text='Count')
        fig_unbox.update_traces(textposition='outside')
        fig_unbox.update_layout(plot_bgcolor='white', paper_bgcolor='white', margin=dict(t=60, b=80), legend=dict(orientation='h', yanchor='top', y=-0.2, xanchor='center', x=0.5, title_text=''))
        st.plotly_chart(fig_unbox, use_container_width=True)
    st.markdown("---")
    tab1, tab2 = st.tabs(["🆕 New Features Detail", "📦 Unboxing Detail"])
    with tab1:
        nf_detail_cols = ['Pillar','Product','Module','Feature','CDL Name','Feature Category','Final Overall Status','Actual Release Date']
        nf_detail_cols_available = [c for c in nf_detail_cols if c in df_rel_nf.columns]
        detail_nf = df_rel_nf[nf_detail_cols_available].reset_index(drop=True)
        st.dataframe(detail_nf, use_container_width=True, column_config=col_config_compact(detail_nf))
        st.download_button("⬇️ Download Released NF Report", export_excel(detail_nf), "report8_released_nf.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with tab2:
        unbox_detail_cols = ['Pillar','Product','Feature','CDL Name','Final Overall Status','Actual Release Date']
        unbox_detail_cols_available = [c for c in unbox_detail_cols if c in df_rel_unbox.columns]
        detail_unbox = df_rel_unbox[unbox_detail_cols_available].reset_index(drop=True)
        st.dataframe(detail_unbox, use_container_width=True, column_config=col_config_compact(detail_unbox))
        st.download_button("⬇️ Download Released Unboxing Report", export_excel(detail_unbox), "report8_released_unboxing.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─────────────────────────────────────────────
# REPORT 9: Feature Overall Status
# ─────────────────────────────────────────────
elif selected_report == "9️⃣  Feature Overall Status":
    st.markdown('<div class="section-header">Feature Count by Final Overall Status — Pillar Wise</div>', unsafe_allow_html=True)
    df_filtered = render_report_filters(df_filtered, 'r9', has_cdl=True)
    ga_base_r9 = pd.Timestamp('2026-07-03'); valid_from_r9 = pd.Timestamp('2020-01-01')
    def get_ga_tier_r9(date, base):
        if pd.isna(date): return None
        if date <= base: return 'GA+0'
        diff = (date - base).days; tier = ((diff - 1) // 7) + 1
        return f'GA+{tier}'
    df_r9 = df_filtered.copy()
    df_r9['Training Required? '] = df_r9['Training Required? '].fillna('Not Set')
    df_r9 = df_r9[df_r9['Training Required? '].isin(['Yes', 'TBD'])].copy()
    df_r9['Final Overall Status'] = df_r9['Final Overall Status'].fillna('Not Set')
    df_r9['CDL Name']             = df_r9['CDL Name'].fillna('Unassigned')
    for col in ['Recording Date', 'Video Ready Date', 'Actual Release Date']:
        if col in df_r9.columns:
            df_r9[col] = pd.to_datetime(df_r9[col], errors='coerce')
            df_r9[col] = df_r9[col].apply(lambda x: x if pd.notna(x) and x >= valid_from_r9 else pd.NaT)
    df_r9['Target GA+']    = df_r9.apply(lambda row: get_ga_tier_r9(row['Recording Date'] + pd.Timedelta(days=4), ga_base_r9) if pd.notna(row['Recording Date']) else ('Recording Dates TBD' if row['Training Required? '] in ['Yes', 'TBD'] else ''), axis=1)
    df_r9['Predicted GA+'] = df_r9.apply(lambda row: get_ga_tier_r9(row['Video Ready Date'] + pd.Timedelta(days=4), ga_base_r9) if pd.notna(row['Video Ready Date']) else ('Recording To Be Submitted' if row['Training Required? '] in ['Yes', 'TBD'] else ''), axis=1)
    df_r9['Actual GA+']    = df_r9.apply(lambda row: get_ga_tier_r9(row['Actual Release Date'], ga_base_r9) if row['Final Overall Status'] == 'Released (A)' and pd.notna(row.get('Actual Release Date')) else '', axis=1)
    df_r9_nf = df_r9[df_r9['Feature Category'] != 'Unboxing'].copy()
    df_r9_ub = df_r9[df_r9['Feature Category'] == 'Unboxing'].copy()
    all_statuses = ['Awaiting CDL Date','Development WIP','In Review (A)','In Review Fixes (A)','In Post Production (A)','In QA (A)','In QA Fixes (A)','In Setup (A)','Released (A)','Feature Dropped','Not Set']
    STATUS_COLORS = {'Awaiting CDL Date':'#94a3b8','Development WIP':'#f59e0b','In Review (A)':'#4f46e5','In Review Fixes (A)':'#7c3aed','In Post Production (A)':'#0ea5e9','In QA (A)':'#10b981','In QA Fixes (A)':'#34d399','In Setup (A)':'#a855f7','Released (A)':'#16a34a','Feature Dropped':'#ef4444','Not Set':'#e2e8f0'}
    def style_r9_detail(row):
        styles = []
        for c in list(row.index):
            if c == 'Target GA+': styles.append('color: #4f46e5; font-weight: bold')
            elif c == 'Predicted GA+': styles.append('color: #f59e0b; font-weight: bold')
            elif c == 'Actual GA+': styles.append('color: #10b981; font-weight: bold')
            else: styles.append('')
        return styles
    def build_pivot(df_src):
        pivot = df_src.groupby(['Pillar','Final Overall Status']).size().unstack(fill_value=0)
        for s in all_statuses:
            if s not in pivot.columns: pivot[s] = 0
        cols_present = [s for s in all_statuses if s in pivot.columns]
        pivot = pivot[cols_present]; pivot['Total'] = pivot.sum(axis=1); pivot = pivot.reset_index()
        return add_total_row(pivot)
    def render_pivot_table(pivot_df):
        num_cols = [c for c in pivot_df.columns if c not in ['Pillar']]
        for c in num_cols:
            try: pivot_df[c] = pivot_df[c].astype(int)
            except: pass
        styled = style_with_total(pivot_df)
        styled = styled.set_properties(**{'text-align': 'left', 'min-width': '60px'})
        styled = styled.set_table_styles([{'selector': 'th', 'props': [('text-align', 'left'), ('white-space', 'nowrap')]}, {'selector': 'td', 'props': [('text-align', 'left'), ('white-space', 'nowrap')]}])
        styled = styled.hide(axis='index')
        st.dataframe(styled, use_container_width=True)
    tab1, tab2 = st.tabs(["🆕 New Features", "📦 Unboxing Videos"])
    with tab1:
        nf_status = df_r9_nf.groupby(['Pillar','Final Overall Status']).size().reset_index(name='Count')
        fig_nf = px.bar(nf_status, x='Pillar', y='Count', color='Final Overall Status', barmode='stack', title='New Features — Final Overall Status by Pillar', color_discrete_map=STATUS_COLORS, text='Count')
        fig_nf.update_traces(textposition='auto', textfont_size=11)
        fig_nf.update_layout(plot_bgcolor='white', paper_bgcolor='white', margin=dict(t=60, b=80), legend=dict(orientation='h', yanchor='top', y=-0.2, xanchor='center', x=0.5, title_text=''))
        st.plotly_chart(fig_nf, use_container_width=True)
        render_pivot_table(build_pivot(df_r9_nf))
        st.markdown("---")
        sel_nf = st.selectbox("Filter by Status", ['All'] + sorted(df_r9_nf['Final Overall Status'].unique().tolist()), key='sel_nf')
        nf_detail_cols = ['Pillar','Product','Module','Feature','CDL Name','Final Overall Status','Feature Category','Target GA+','Predicted GA+','Actual GA+']
        nf_detail_cols_available = [c for c in nf_detail_cols if c in df_r9_nf.columns]
        det_nf = df_r9_nf[nf_detail_cols_available].copy()
        if sel_nf != 'All': det_nf = det_nf[det_nf['Final Overall Status'] == sel_nf]
        st.markdown('<p style="font-size:0.82rem;"><span style="color:#4f46e5; font-weight:bold;">🔵 Target GA+</span> = Recording Date + 4 days &nbsp;|&nbsp; <span style="color:#f59e0b; font-weight:bold;">🟠 Predicted GA+</span> = Video Ready Date + 4 days &nbsp;|&nbsp; <span style="color:#10b981; font-weight:bold;">🟢 Actual GA+</span> = Actual Release Date (Released only)</p>', unsafe_allow_html=True)
        st.dataframe(det_nf.style.apply(style_r9_detail, axis=1).hide(axis='index'), use_container_width=True, column_config=col_config_compact(det_nf))
        st.download_button("⬇️ Download New Features Status", export_excel(det_nf), "report9_nf_status.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key='dl_nf_status')
    with tab2:
        if len(df_r9_ub) == 0:
            st.info("No Unboxing videos found.")
        else:
            ub_status = df_r9_ub.groupby(['Pillar','Final Overall Status']).size().reset_index(name='Count')
            fig_ub = px.bar(ub_status, x='Pillar', y='Count', color='Final Overall Status', barmode='stack', title='Unboxing Videos — Final Overall Status by Pillar', color_discrete_map=STATUS_COLORS, text='Count')
            fig_ub.update_traces(textposition='auto', textfont_size=11)
            fig_ub.update_layout(plot_bgcolor='white', paper_bgcolor='white', margin=dict(t=60, b=80), legend=dict(orientation='h', yanchor='top', y=-0.2, xanchor='center', x=0.5, title_text=''))
            st.plotly_chart(fig_ub, use_container_width=True)
            render_pivot_table(build_pivot(df_r9_ub))
            st.markdown("---")
            sel_ub = st.selectbox("Filter by Status", ['All'] + sorted(df_r9_ub['Final Overall Status'].unique().tolist()), key='sel_ub')
            ub_detail_cols = ['Pillar','Product','Feature','CDL Name','Final Overall Status','Target GA+','Predicted GA+','Actual GA+']
            ub_detail_cols_available = [c for c in ub_detail_cols if c in df_r9_ub.columns]
            det_ub = df_r9_ub[ub_detail_cols_available].copy()
            if sel_ub != 'All': det_ub = det_ub[det_ub['Final Overall Status'] == sel_ub]
            st.markdown('<p style="font-size:0.82rem;"><span style="color:#4f46e5; font-weight:bold;">🔵 Target GA+</span> = Recording Date + 4 days &nbsp;|&nbsp; <span style="color:#f59e0b; font-weight:bold;">🟠 Predicted GA+</span> = Video Ready Date + 4 days &nbsp;|&nbsp; <span style="color:#10b981; font-weight:bold;">🟢 Actual GA+</span> = Actual Release Date (Released only)</p>', unsafe_allow_html=True)
            st.dataframe(det_ub.style.apply(style_r9_detail, axis=1).hide(axis='index'), use_container_width=True, column_config=col_config_compact(det_ub))
            st.download_button("⬇️ Download Unboxing Status", export_excel(det_ub), "report9_ub_status.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key='dl_ub_status')

# ─────────────────────────────────────────────
# REPORT 10
# ─────────────────────────────────────────────
elif selected_report == "🔟  Released Course List":
    st.markdown('<div class="section-header">Released Course List — New Features & Unboxing</div>', unsafe_allow_html=True)
    df_filtered = render_report_filters(df_filtered, 'r10', has_cdl=True)
    df_r10       = df_filtered[df_filtered['Final Overall Status'] == 'Released (A)'].copy()
    df_r10_nf    = df_r10[df_r10['Feature Category'] != 'Unboxing']
    df_r10_unbox = df_r10[df_r10['Feature Category'] == 'Unboxing']
    col1, col2 = st.columns(2)
    col1.markdown(metric_card(len(df_r10_nf), "Released New Features"), unsafe_allow_html=True)
    col2.markdown(metric_card(len(df_r10_unbox), "Released Unboxing Videos"), unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("---")
    col_map = {'Pillar':'Pillar','Product':'Product','Module':'Module','Feature Category':'Feature Category','Tags ( AI / Redwood/AI Agent)':'Tag','DCFR Ranking':'DCFR Ranking','CDL Name':'CDL','OMBP Mapping':'OMBP','Feature':'New Feature Name','Actual Release Date':'Release Date','MyLearn Link to Video':'Link to Video','Container ID':'Course Container ID','Component ID':'Course Component ID'}
    available_cols  = [c for c in col_map.keys() if c in df_r10.columns]
    df_r10_display  = df_r10[available_cols].copy().rename(columns=col_map).reset_index(drop=True)
    filter_cat = st.selectbox("Filter by Category", ['All','New Features','Unboxing'], key='sel_r10_cat')
    if filter_cat == 'New Features': df_r10_display = df_r10_display[df_r10_display['Feature Category'] != 'Unboxing']
    elif filter_cat == 'Unboxing': df_r10_display = df_r10_display[df_r10_display['Feature Category'] == 'Unboxing']
    pillar_filter = st.multiselect("Filter by Pillar", options=sorted(df_r10_display['Pillar'].dropna().unique().tolist()), default=sorted(df_r10_display['Pillar'].dropna().unique().tolist()), key='sel_r10_pillar')
    if pillar_filter: df_r10_display = df_r10_display[df_r10_display['Pillar'].isin(pillar_filter)]
    st.dataframe(df_r10_display, use_container_width=True, column_config=col_config_compact(df_r10_display))
    st.download_button("⬇️ Download Released Course List", export_excel(df_r10_display), "report10_released_course_list.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─────────────────────────────────────────────
# REPORT 11: Issues Tracker
# ─────────────────────────────────────────────
elif selected_report == "1️⃣1️⃣  Issues Tracker — Pillar & CDL Wise":
    st.markdown('<div class="section-header">Issues Tracker — Pillar & CDL Wise</div>', unsafe_allow_html=True)
    df_filtered = render_report_filters(df_filtered, 'r11', has_cdl=True)
    df_r11 = df_filtered.copy()
    df_r11['CDL Name']       = df_r11['CDL Name'].fillna('Unassigned')
    df_r11['Has Issue']      = df_r11['Type of Issue'].notna() & (df_r11['Type of Issue'] != '')
    df_r11['Issue Resolved'] = df_r11['Issue Resolution Outcome'].notna() & (df_r11['Issue Resolution Outcome'] != '')
    df_issues = df_r11[df_r11['Has Issue']].copy()

    col1, col2, col3 = st.columns(3)
    col1.markdown(metric_card(len(df_issues), "Total Features with Issues"), unsafe_allow_html=True)
    col2.markdown(metric_card(len(df_issues[df_issues['Issue Resolved']]), "✅ Resolved"), unsafe_allow_html=True)
    col3.markdown(metric_card(len(df_issues[~df_issues['Issue Resolved']]), "⚠️ Unresolved"), unsafe_allow_html=True)

    if len(df_issues) == 0:
        st.success("🎉 No issues found!")
    else:
        col1, col2 = st.columns(2)
        with col1:
            pillar_issues = df_issues.groupby('Pillar').size().reset_index(name='Issues')
            st.plotly_chart(color_bar_chart(pillar_issues, 'Pillar', 'Issues', 'Issues by Pillar', 'Pillar'), use_container_width=True)
        with col2:
            cdl_issues = df_issues.groupby(['CDL Name', 'Issue Resolved']).size().reset_index(name='Count')
            cdl_issues['Status'] = cdl_issues['Issue Resolved'].map({True: 'Resolved', False: 'Unresolved'})
            fig2 = px.bar(cdl_issues, x='CDL Name', y='Count', color='Status',
                          title='Issues by CDL — Resolved vs Unresolved',
                          color_discrete_map={'Resolved': '#10b981', 'Unresolved': '#ef4444'},
                          text='Count', barmode='group')
            fig2.update_traces(textposition='outside')
            fig2.update_layout(plot_bgcolor='white', paper_bgcolor='white', xaxis_tickangle=-30)
            st.plotly_chart(fig2, use_container_width=True)
        detail = df_issues[['Pillar', 'Product', 'Feature', 'CDL Name', 'Type of Issue', 'Issue Resolution Outcome', 'Details of the Issue faced ']].reset_index(drop=True)
        st.dataframe(detail, use_container_width=True, column_config=col_config_compact(detail))
        st.download_button("⬇️ Download Issues Report", export_excel(detail), "report11_issues.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.markdown("---")
    st.markdown('<div class="section-header">🔴 Features Requiring D01 / D02 Pod</div>', unsafe_allow_html=True)

    df_pod = df_r11[df_r11['Issue Resolution Outcome'].fillna('').str.strip().str.lower().isin(
        ['required d01 pod', 'requires d01 pod', 'required d02 pod', 'requires d02 pod']
    )].copy()

    if len(df_pod) == 0:
        st.info("No features requiring D01 or D02 Pod found.")
    else:
        pod_cols = ['Pillar', 'Module', 'Product', 'Feature', 'Tags ( AI / Redwood/AI Agent)', 'DCFR Ranking', 'CDL Name', 'Issue Resolution Outcome']
        pod_cols_available = [c for c in pod_cols if c in df_pod.columns]
        df_pod_display = df_pod[pod_cols_available].copy().reset_index(drop=True)
        df_pod_display = df_pod_display.rename(columns={'Tags ( AI / Redwood/AI Agent)': 'Feature Tag'})

        def style_pod(row):
            val = str(row.get('Issue Resolution Outcome', '')).strip().lower()
            if 'd01' in val:
                return ['background-color: #fff3cd; color: #856404; font-weight: bold'] * len(row)
            elif 'd02' in val:
                return ['background-color: #fce4d6; color: #9c3000; font-weight: bold'] * len(row)
            return [''] * len(row)

        st.markdown(f"**{len(df_pod)} feature(s)** require a D01 or D02 Pod session.")
        st.dataframe(
            df_pod_display.style.apply(style_pod, axis=1).hide(axis='index'),
            use_container_width=True,
            column_config=col_config_compact(df_pod_display)
        )
        st.markdown("""<div style="background-color:#fff8f0; padding:8px 14px; border-radius:6px;
            border-left:4px solid #f59e0b; margin-top:6px; font-size:0.82rem; color:#444;">
            <span style="background:#fff3cd; padding:2px 6px; border-radius:3px; font-weight:bold; color:#856404;">🟡 Requires D01 Pod</span> &nbsp;|&nbsp;
            <span style="background:#fce4d6; padding:2px 6px; border-radius:3px; font-weight:bold; color:#9c3000;">🟠 Requires D02 Pod</span>
        </div>""", unsafe_allow_html=True)
        st.download_button("⬇️ Download D01/D02 Pod Features", export_excel(df_pod_display), "report11_pod_features.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_pod")
# ─────────────────────────────────────────────
# REPORT 12: GA+ Release
# ─────────────────────────────────────────────
elif selected_report == "1️⃣2️⃣  GA+ Release":
    st.markdown('<div class="section-header">GA+ Release — Target vs Expected vs Actual Release Tier</div>', unsafe_allow_html=True)
    df_filtered = render_report_filters(df_filtered, 'r12', has_cdl=True)
    default_base = datetime(2026, 7, 3)
    ga_base_date = st.date_input("Select First Friday of Release Month (GA+0 cutoff date)", value=default_base, key='ga_base_date')
    ga_base_ts   = pd.Timestamp(ga_base_date)
    st.info(f"ℹ️ GA+0 cutoff = {ga_base_ts.strftime('%d-%b-%Y')} | Each subsequent Friday = next GA+ tier")
    def get_ga_tier(date, base_date):
        if pd.isna(date): return 'Not Set'
        if date <= base_date: return f'GA+0\n({base_date.strftime("%d-%b-%Y")})'
        diff = (date - base_date).days; tier = ((diff - 1) // 7) + 1
        return f'GA+{tier}\n({(base_date + pd.Timedelta(weeks=tier)).strftime("%d-%b-%Y")})'
    def get_ga_tier_num(tier_str):
        if tier_str == 'Not Set': return 999
        try: return int(tier_str.split('\n')[0].replace('GA+', ''))
        except: return 999
    df_r12 = df_filtered.copy()
    valid_from = pd.Timestamp('2020-01-01')
    for col in ['Recording Date', 'Video Ready Date', 'Actual Release Date']:
        df_r12[col] = pd.to_datetime(df_r12[col], errors='coerce').apply(lambda x: x if pd.notna(x) and x >= valid_from else pd.NaT)
    df_r12['Training Required? '] = df_r12['Training Required? '].fillna('Not Set')
    df_r12_nf    = df_r12[(df_r12['Feature Category'] != 'Unboxing') & (df_r12['Training Required? '] == 'Yes')].copy()
    df_r12_unbox = df_r12[(df_r12['Feature Category'] == 'Unboxing') & (df_r12['Training Required? '] == 'Yes')].copy()
    df_r12_combined = pd.concat([df_r12_nf, df_r12_unbox], ignore_index=True)
    def map_pillar_group_r12(pillar):
        if str(pillar).startswith('CX'): return 'CX'
        return pillar
    df_r12_combined['Pillar Group'] = df_r12_combined['Pillar'].apply(map_pillar_group_r12)
    df_r12_combined['Target Release Date Calc']   = df_r12_combined['Recording Date'].apply(lambda x: x + pd.Timedelta(days=4) if pd.notna(x) else pd.NaT)
    df_r12_combined['Expected Release Date Calc'] = df_r12_combined['Video Ready Date'].apply(lambda x: x + pd.Timedelta(days=4) if pd.notna(x) else pd.NaT)
    df_r12_combined['Actual Released Date Calc']  = df_r12_combined.apply(lambda row: row['Actual Release Date'] if row['Final Overall Status'] == 'Released (A)' else pd.NaT, axis=1)
    df_r12_combined['Target Release Tier']   = df_r12_combined['Target Release Date Calc'].apply(lambda x: get_ga_tier(x, ga_base_ts))
    df_r12_combined['Expected Release Tier'] = df_r12_combined['Expected Release Date Calc'].apply(lambda x: get_ga_tier(x, ga_base_ts))
    df_r12_combined['Actual Released Tier']  = df_r12_combined.apply(lambda row: get_ga_tier(row['Actual Released Date Calc'] if pd.notna(row['Actual Released Date Calc']) else (row['Target Release Date Calc'] if row['Final Overall Status'] == 'Released (A)' and pd.notna(row['Target Release Date Calc']) else pd.NaT), ga_base_ts), axis=1)
    for col, src in [('Target Release Date Display','Target Release Date Calc'),('Expected Release Date Display','Expected Release Date Calc'),('Actual Released Date Display','Actual Released Date Calc'),('Recording Date Display','Recording Date'),('Video Ready Date Display','Video Ready Date')]:
        df_r12_combined[col] = df_r12_combined[src].apply(lambda x: x.strftime('%d-%b-%Y') if pd.notna(x) else 'Not Set')
    all_tiers_list = []
    for col in ['Target Release Tier','Expected Release Tier','Actual Released Tier']:
        all_tiers_list += [t for t in df_r12_combined[col].unique() if t != 'Not Set']
    all_tiers = sorted(list(dict.fromkeys(all_tiers_list)), key=get_ga_tier_num)
    # Not Set intentionally excluded from summary table
    st.markdown('<div class="section-header">🎯 Overall GA+ Release Summary — All Features</div>', unsafe_allow_html=True)

    # ── CHANGE 2: Two % columns — % of Target + Release % ──
    total_training_yes = len(df_r12_combined)  # already Training=Yes NF + Unboxing
    summary_box_rows = []
    for tier in all_tiers:
        target_count   = len(df_r12_combined[df_r12_combined['Target Release Tier'] == tier])
        expected_count = len(df_r12_combined[df_r12_combined['Expected Release Tier'] == tier])
        actual_count   = len(df_r12_combined[df_r12_combined['Actual Released Tier'] == tier])
        pct_of_target  = round(actual_count / target_count * 100, 1) if target_count > 0 else 0.0
        release_pct    = round(actual_count / total_training_yes * 100, 1) if total_training_yes > 0 else 0.0
        summary_box_rows.append({
            'GA+ Tier': tier,
            'Target Count': target_count,
            'Expected to be Released': expected_count,
            'Actual Released': actual_count,
            '% of Target': f'{pct_of_target}%' if tier != 'Not Set' else '—',
            'Release %': f'{release_pct}%' if tier != 'Not Set' else '—',
        })
    summary_box_df = pd.DataFrame(summary_box_rows)
    total_target   = summary_box_df['Target Count'].sum()
    total_expected = summary_box_df['Expected to be Released'].sum()
    total_actual   = summary_box_df['Actual Released'].sum()
    total_pct_of_target = round(total_actual / total_target * 100, 1) if total_target > 0 else 0.0
    total_release_pct   = round(total_actual / total_training_yes * 100, 1) if total_training_yes > 0 else 0.0
    total_sb = {'GA+ Tier': 'Total', 'Target Count': total_target, 'Expected to be Released': total_expected,
                'Actual Released': total_actual, '% of Target': f'{total_pct_of_target}%', 'Release %': f'{total_release_pct}%'}
    summary_box_with_total = pd.concat([summary_box_df, pd.DataFrame([total_sb])], ignore_index=True)
    def style_summary_box(row):
        if row['GA+ Tier'] == 'Total': return ['font-weight: bold; background-color: #f0f0f0'] * len(row)
        styles = []
        for c in list(row.index):
            if c == 'Target Count': styles.append('color: #4f46e5; font-weight: bold')
            elif c == 'Expected to be Released': styles.append('color: #f59e0b; font-weight: bold')
            elif c == 'Actual Released': styles.append('color: #10b981; font-weight: bold')
            elif c == '% of Target': styles.append('color: #4f46e5; font-weight: bold')
            elif c == 'Release %': styles.append('color: #10b981; font-weight: bold')
            elif c == 'GA+ Tier': styles.append('font-weight: bold')
            else: styles.append('')
        return styles
    styled_summary_box = summary_box_with_total.style.apply(style_summary_box, axis=1).hide(axis='index')
    st.dataframe(styled_summary_box, use_container_width=True, column_config=col_config_compact(summary_box_with_total))
    # Count Not Set features for the note
    not_set_count = len(df_r12_combined[df_r12_combined['Target Release Tier'] == 'Not Set'])
    st.markdown(f"""
    <div style="background-color:#fff3cd;padding:10px 16px;border-radius:6px;border-left:4px solid #f59e0b;margin-top:6px;font-size:0.8rem;color:#555;">
        <b>📌 Note:</b> <b>{not_set_count} feature(s) with "Not Set" Recording Date are excluded from this table.</b><br>
        <span style="color:#777;">"Not Set" means the feature has no Recording Date entered yet, so no GA+ release tier can be calculated. 
        These features are still tracked in the detail table below and in other reports.</span>
    </div>
    """, unsafe_allow_html=True)
    st.markdown(f"""<div style="background-color:#fffbeb;padding:12px 16px;border-radius:6px;border-left:4px solid #f59e0b;margin-top:8px;font-size:0.82rem;color:#444;">
    <b>📌 How to read this table:</b><br><br>
    &bull; <b style="color:#4f46e5;">Target Count</b> — Features targeted for that GA+ tier based on <b>Recording Date + 4 days</b>.<br><br>
    &bull; <b style="color:#f59e0b;">Expected to be Released</b> — Based on <b>Video Ready Date + 4 days</b>.<br><br>
    &bull; <b style="color:#10b981;">Actual Released</b> — Features confirmed as <b>Released (A)</b> in that tier.<br><br>
    &bull; <b style="color:#4f46e5;">% of Target</b> — <b>Actual Released ÷ Target Count</b> for that tier — how well we hit each tier's target.<br><br>
    &bull; <b style="color:#10b981;">Release %</b> — <b>Actual Released ÷ Total Training=Yes ({total_training_yes})</b> — overall progress. Includes NF (Training=Yes) + Unboxing (Training=Yes). Overall: <b>{total_release_pct}%</b><br><br>
    &bull; <b>Note:</b> Features with Training Required = Not Set are excluded from the total denominator.
    </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown('<div class="section-header">📦 Pillar Wise — GA+ Release Summary</div>', unsafe_allow_html=True)
    pillar_groups = ['CX', 'ERP', 'HCM', 'PRC', 'SCM']
    box_cols = st.columns(5)
    for col, pg in zip(box_cols, pillar_groups):
        pg_df    = df_r12_combined[df_r12_combined['Pillar Group'] == pg]
        nf_count = len(pg_df[pg_df['Feature Category'] != 'Unboxing'])
        ub_count = len(pg_df[pg_df['Feature Category'] == 'Unboxing'])
        tier_lines = ''
        for tier in all_tiers:
            t_count = len(pg_df[pg_df['Target Release Tier'] == tier])
            e_count = len(pg_df[pg_df['Expected Release Tier'] == tier])
            a_count = len(pg_df[pg_df['Actual Released Tier'] == tier])
            tier_label = tier.replace('\n', '<br>')
            tier_lines += f'<div style="font-size:0.72rem; margin-bottom:0.3rem;"><b>{tier_label}</b>: <span style="color:#4f46e5">T:{t_count}</span> | <span style="color:#f59e0b">E:{e_count}</span> | <span style="color:#10b981">A:{a_count}</span></div>'
        col.markdown(f'<div class="metric-card" style="border-left-color:#4f46e5; text-align:left; padding:1rem;"><div style="font-size:1rem; font-weight:700; color:#1a1a2e; margin-bottom:0.3rem;">{pg}</div><div style="font-size:0.75rem; color:#6c757d; margin-bottom:0.4rem;">(NF: {nf_count} | Unboxing: {ub_count})</div>{tier_lines}</div>', unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown('<div class="section-header">📊 GA+ Tier Summary — Pillar Wise</div>', unsafe_allow_html=True)
    summary_rows = []
    for pg in pillar_groups:
        pg_df = df_r12_combined[df_r12_combined['Pillar Group'] == pg]
        row = {'Pillar': pg}
        for tier in all_tiers:
            row[f'{tier} Target']   = len(pg_df[pg_df['Target Release Tier'] == tier])
            row[f'{tier} Expected'] = len(pg_df[pg_df['Expected Release Tier'] == tier])
            row[f'{tier} Actual']   = len(pg_df[pg_df['Actual Released Tier'] == tier])
        summary_rows.append(row)
    summary_df = pd.DataFrame(summary_rows)
    summary_with_total = add_total_row(summary_df)
    def style_ga_summary(row):
        if row['Pillar'] == 'Total': return ['font-weight: bold; background-color: #f0f0f0'] * len(row)
        styles = []
        for c in list(row.index):
            if 'Target' in str(c): styles.append('color: #4f46e5; font-weight: bold')
            elif 'Expected' in str(c): styles.append('color: #f59e0b; font-weight: bold')
            elif 'Actual' in str(c): styles.append('color: #10b981; font-weight: bold')
            else: styles.append('')
        return styles
    st.dataframe(summary_with_total.style.apply(style_ga_summary, axis=1).hide(axis='index'), use_container_width=True, column_config=col_config_compact(summary_with_total))
    st.markdown("---")
    col_f1, col_f2 = st.columns(2)
    with col_f1: tier_filter = st.selectbox("Filter by Target Tier", ['All'] + all_tiers, key='ga_tier_filter')
    with col_f2: cat_filter  = st.selectbox("Filter by Category", ['All','New Features','Unboxing'], key='ga_cat_filter')
    detail_cols = ['Pillar','Product','Module','Feature','Feature Category','Tags ( AI / Redwood/AI Agent)','DCFR Ranking','CDL Name','Recording Date Display','Target Release Date Display','Target Release Tier','Video Ready Date Display','Expected Release Date Display','Expected Release Tier','Actual Released Date Display','Actual Released Tier']
    detail_cols_available = [c for c in detail_cols if c in df_r12_combined.columns]
    detail_df = df_r12_combined[detail_cols_available].copy()
    detail_df = detail_df.rename(columns={'Tags ( AI / Redwood/AI Agent)':'Tags','Recording Date Display':'Recording Date','Target Release Date Display':'Target Release Date','Video Ready Date Display':'Video Ready Date','Expected Release Date Display':'Expected Release Date','Actual Released Date Display':'Actual Released Date'})
    if tier_filter != 'All': detail_df = detail_df[detail_df['Target Release Tier'] == tier_filter]
    if cat_filter == 'New Features': detail_df = detail_df[detail_df['Feature Category'] != 'Unboxing']
    elif cat_filter == 'Unboxing': detail_df = detail_df[detail_df['Feature Category'] == 'Unboxing']
    def style_ga_detail(row):
        styles = []
        for c in list(row.index):
            if c in ['Target Release Date','Target Release Tier']: styles.append('color: #4f46e5; font-weight: bold')
            elif c in ['Expected Release Date','Expected Release Tier']: styles.append('color: #f59e0b; font-weight: bold')
            elif c in ['Actual Released Date','Actual Released Tier']: styles.append('color: #10b981; font-weight: bold')
            else: styles.append('')
        return styles
    st.dataframe(detail_df.style.apply(style_ga_detail, axis=1).hide(axis='index'), use_container_width=True, column_config=col_config_compact(detail_df))
    st.download_button("⬇️ Download GA+ Release Report", export_excel(detail_df), "report12_ga_release.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─────────────────────────────────────────────
# REPORT 13: 26C New Features Release Dashboard
# ─────────────────────────────────────────────
elif selected_report == "1️⃣3️⃣  26C New Features Release Dashboard":
    st.markdown('<div class="section-header">26C New Features Release Dashboard — Executive Summary</div>', unsafe_allow_html=True)
    df_filtered = render_report_filters(df_filtered, 'r13', has_cdl=False)
    default_base_r13 = datetime(2026, 7, 3)
    ga_base_date_r13 = st.date_input("Select First Friday of Release Month (GA+0 cutoff date)", value=default_base_r13, key='ga_base_date_r13')
    ga_base_ts_r13   = pd.Timestamp(ga_base_date_r13)
    st.info(f"ℹ️ GA+0 cutoff = {ga_base_ts_r13.strftime('%d-%b-%Y')} | Each subsequent Friday = next GA+ tier")
    def get_ga_tier_r13(date, base_date):
        if pd.isna(date): return 'Not Set'
        if date <= base_date: return 'GA+0'
        diff = (date - base_date).days; tier = ((diff - 1) // 7) + 1
        return f'GA+{tier}'
    df_r13 = df_filtered.copy()
    valid_from = pd.Timestamp('2020-01-01')
    for col in ['Recording Date', 'Video Ready Date', 'Actual Release Date']:
        df_r13[col] = pd.to_datetime(df_r13[col], errors='coerce').apply(lambda x: x if pd.notna(x) and x >= valid_from else pd.NaT)
    df_r13['Training Required? '] = df_r13['Training Required? '].fillna('Not Set')
    def map_pg_r13(pillar):
        if str(pillar).startswith('CX'): return 'CX'
        return pillar
    df_r13['Pillar Group'] = df_r13['Pillar'].apply(map_pg_r13)
    df_r13_nf    = df_r13[(df_r13['Feature Category'] != 'Unboxing') & (df_r13['Training Required? '] == 'Yes')].copy()
    df_r13_unbox = df_r13[(df_r13['Feature Category'] == 'Unboxing') & (df_r13['Training Required? '] == 'Yes')].copy()
    df_r13_all   = pd.concat([df_r13_nf, df_r13_unbox], ignore_index=True)
    df_r13_all['Target Tier'] = df_r13_all['Recording Date'].apply(lambda x: get_ga_tier_r13(x + pd.Timedelta(days=4), ga_base_ts_r13) if pd.notna(x) else 'Not Set')
    # Build lookup: Combined Primary NF → Actual Release Date
    primary_release_map = {}
    for _, row in df_r13_all[df_r13_all['Feature Classification'] == 'Combined Primary NF'].iterrows():
        cfvn = row.get('Combined Feature Video Name')
        ard  = row.get('Actual Release Date')
        if pd.notna(cfvn) and str(cfvn).strip() and pd.notna(ard):
            primary_release_map[str(cfvn).strip()] = ard
    def get_released_tier_r13(row):
        status = row['Final Overall Status']
        ard    = row['Actual Release Date']
        fc     = str(row.get('Feature Classification', '') or '').strip()
        cfvn   = str(row.get('Combined Feature Video Name', '') or '').strip()
        if status != 'Released (A)':
            return 'Not Set'
        # For Combined Bundled NF — inherit release date from Primary NF
        if fc == 'Combined Bundled NF' and pd.isna(ard) and cfvn in primary_release_map:
            ard = primary_release_map[cfvn]
        if pd.notna(ard):
            return get_ga_tier_r13(ard, ga_base_ts_r13)
        return 'Not Set'
    df_r13_all['Released Tier'] = df_r13_all.apply(get_released_tier_r13, axis=1)
    all_tier_nums = set()
    for t in list(df_r13_all['Target Tier'].unique()) + list(df_r13_all['Released Tier'].unique()):
        if t != 'Not Set':
            try: all_tier_nums.add(int(t.replace('GA+', '')))
            except: pass
    sorted_tier_nums = sorted(all_tier_nums)
    def tier_date_label(tier_num, base):
        d = base + pd.Timedelta(weeks=tier_num)
        return f"GA+{tier_num} ({d.strftime('%d-%b-%Y')})"
    pillar_groups_r13 = ['HCM', 'ERP', 'SCM', 'PRC', 'CX']
    table_rows = []
    for pg in pillar_groups_r13:
        pg_df    = df_r13_all[df_r13_all['Pillar Group'] == pg]
        pg_nf    = df_r13_nf[df_r13_nf['Pillar Group'] == pg]
        pg_unbox = df_r13_unbox[df_r13_unbox['Pillar Group'] == pg]
        nf_count    = len(pg_nf)
        unbox_count = len(pg_unbox)
        total    = nf_count + unbox_count
        released = len(pg_df[pg_df['Final Overall Status'] == 'Released (A)'])
        pct_rel  = round(released / total * 100, 1) if total > 0 else 0.0
        row = {
            'Pillar': pg,
            '# of NF (Y)': nf_count,
            '# of Unboxing (Y)': unbox_count,
            'Total': total,
            'Total Released': released,
            '% Released': f'{pct_rel}%',
        }
        for tn in sorted_tier_nums:
            tier_str = f'GA+{tn}'; label_p = tier_date_label(tn, ga_base_ts_r13)
            row[f'{label_p} Planned']  = len(pg_df[pg_df['Target Tier'] == tier_str])
            row[f'{label_p} Released'] = len(pg_df[pg_df['Released Tier'] == tier_str])
        table_rows.append(row)
    total_row = {'Pillar': 'Total'}
    for col in table_rows[0].keys():
        if col not in ['Pillar', '% Released']:
            total_row[col] = sum(r[col] for r in table_rows if isinstance(r[col], (int, float)))
        elif col == '% Released':
            total_row[col] = ''
    total_row['% Released'] = f"{round(total_row['Total Released'] / total_row['Total'] * 100, 1)}%" if total_row['Total'] > 0 else '0%'
    table_rows.append(total_row)
    df_table = pd.DataFrame(table_rows)
    total_videos_grand = total_row['Total']
    pct_row = {'Pillar': '% Released of Total', '# of NF (Y)': '', '# of Unboxing (Y)': '', 'Total': '', '% Released': ''}
    pct_row['Total Released'] = f"{round(total_row['Total Released'] / total_videos_grand * 100)}%" if total_videos_grand > 0 else '0%'
    for tn in sorted_tier_nums:
        label_p = tier_date_label(tn, ga_base_ts_r13)
        pct_row[f'{label_p} Planned']  = f"{round(total_row.get(f'{label_p} Planned', 0) / total_videos_grand * 100)}%" if total_videos_grand > 0 else '0%'
        pct_row[f'{label_p} Released'] = f"{round(total_row.get(f'{label_p} Released', 0) / total_videos_grand * 100)}%" if total_videos_grand > 0 else '0%'
    df_final = pd.concat([df_table, pd.DataFrame([pct_row])], ignore_index=True)
    st.markdown('<div class="section-header">📊 Executive Summary Table</div>', unsafe_allow_html=True)
    st.markdown(
        '<div style="background-color:#e8f4f8; padding:8px 14px; border-radius:6px; '
        'border-left:4px solid #0078d4; margin-bottom:8px; font-size:0.85rem; color:#0c5460;">'
        'ℹ️ <b># of NF (Y)</b> = NF with Training=Yes &nbsp;|&nbsp; '
        '<b># of Unboxing (Y)</b> = Unboxing with Training=Yes &nbsp;|&nbsp; '
        '<b>Total</b> = NF(Yes) + Unboxing(Yes) only — features with Training=Not Set are excluded.'
        '</div>',
        unsafe_allow_html=True
    )
    def style_exec_table(row):
        styles = []
        for c in list(row.index):
            if row['Pillar'] == 'Total': styles.append('font-weight: bold; background-color: #f0f0f0')
            elif row['Pillar'] == '% Released of Total': styles.append('font-weight: bold; color: #4f46e5; background-color: #e8eaf6')
            elif c == 'Total Released': styles.append('background-color: #d4edda; font-weight: bold; color: #155724')
            elif c == '% Released': styles.append('background-color: #e8f5e9; font-weight: bold; color: #155724')
            elif 'Released' in str(c) and c not in ['Total Released', '% Released']: styles.append('background-color: #fff3cd; color: #856404')
            elif 'Planned' in str(c): styles.append('background-color: #e8f4f8; color: #0c5460')
            else: styles.append('')
        return styles
    st.dataframe(df_final.style.apply(style_exec_table, axis=1).hide(axis='index'), use_container_width=True, column_config=col_config_compact(df_final))
    st.markdown("""<div style="background-color:#fffbeb; padding:10px 14px; border-radius:6px; border-left:4px solid #f59e0b; margin-top:8px; font-size:0.82rem; color:#444;">
    📌 <b>Notes:</b><br>
    &bull; <b># of NF (Y)</b> = New Features where Training Required = <b>Yes</b> only.<br>
    &bull; <b># of Unboxing (Y)</b> = Unboxing videos where Training Required = <b>Yes</b> only.<br>
    &bull; <b>Total</b> = NF (Yes) + Unboxing (Yes). Features with Training Required = <b>Not Set / TBD / No</b> are <b>excluded</b> from this total.<br>
    &bull; <b>% Released</b> = Total Released ÷ Total (Training=Yes features only) per pillar.
    </div>""", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown('<div class="section-header">⬇️ Download for Email</div>', unsafe_allow_html=True)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    def generate_exec_excel(df_data, base_ts, tier_nums):
        wb = Workbook(); ws = wb.active; ws.title = "26C Release Dashboard"
        thin = Side(style='thin', color='AAAAAA'); border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells(f'A1:{get_column_letter(len(df_data.columns))}1')
        ws['A1'] = '26C New Features Release Dashboard'
        ws['A1'].font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
        ws['A1'].fill = PatternFill('solid', start_color='1F3864')
        ws['A1'].alignment = center; ws.row_dimensions[1].height = 25
        hdr_fill = PatternFill('solid', start_color='2E75B6')
        hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        for c_idx, col in enumerate(df_data.columns, 1):
            cell = ws.cell(row=2, column=c_idx, value=col)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = border
        ws.row_dimensions[2].height = 40
        green_fill = PatternFill('solid', start_color='C6EFCE'); pct_rel_fill = PatternFill('solid', start_color='E8F5E9')
        planned_fill = PatternFill('solid', start_color='DDEBF7'); rel_fill = PatternFill('solid', start_color='FFF2CC')
        total_fill = PatternFill('solid', start_color='D9D9D9'); pct_fill = PatternFill('solid', start_color='E8EAF6')
        for r_idx, row in enumerate(df_data.itertuples(index=False), 3):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.alignment = center; cell.border = border; cell.font = Font(name='Arial', size=9)
                col_name = df_data.columns[c_idx - 1]
                if str(getattr(row, 'Pillar', '')) == 'Total':
                    cell.fill = total_fill; cell.font = Font(name='Arial', bold=True, size=9)
                elif str(getattr(row, 'Pillar', '')) == '% Released of Total':
                    cell.fill = pct_fill; cell.font = Font(name='Arial', bold=True, size=9, color='1F3864')
                elif col_name == 'Total Released': cell.fill = green_fill
                elif col_name == '% Released': cell.fill = pct_rel_fill; cell.font = Font(name='Arial', bold=True, size=9, color='155724')
                elif 'Released' in str(col_name) and col_name not in ['Total Released','% Released']: cell.fill = rel_fill
                elif 'Planned' in str(col_name): cell.fill = planned_fill
        for c_idx, col in enumerate(df_data.columns, 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = 18 if c_idx > 6 else 14
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()
    excel_data = generate_exec_excel(df_final, ga_base_ts_r13, sorted_tier_nums)
    st.download_button("⬇️ Download 26C Release Dashboard (Excel)", excel_data, "26C_Release_Dashboard.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
# ─────────────────────────────────────────────
# REPORT 14: Daily CDL Status Report
# ─────────────────────────────────────────────
elif selected_report == "1️⃣4️⃣  Daily CDL Status Report":
    import numpy as np
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    st.markdown('<div class="section-header">📋 Daily CDL Status Report</div>', unsafe_allow_html=True)
    df_filtered = render_report_filters(df_filtered, 'r14', has_cdl=True)
    default_base_r14 = datetime(2026, 7, 3)
    ga_base_r14    = st.date_input("Select GA+0 Base Date (First Friday of Release Month)", value=default_base_r14, key='ga_base_r14')
    ga_base_ts_r14 = pd.Timestamp(ga_base_r14)
    today_r14      = pd.Timestamp(datetime.today().date())
    st.info(f"ℹ️ Report generated as of: {today_r14.strftime('%d-%b-%Y')} | GA+0 = {ga_base_ts_r14.strftime('%d-%b-%Y')}")
    def get_ga_tier_r14(date, base):
        if pd.isna(date): return 'Not Set'
        if date <= base: return 'GA+0'
        diff = (date - base).days; tier = ((diff - 1) // 7) + 1
        return f'GA+{tier}'
    df_r14 = df_filtered.copy()
    valid_from = pd.Timestamp('2020-01-01')
    for col in ['Recording Date', 'Video Ready Date', 'Actual Release Date']:
        if col in df_r14.columns:
            df_r14[col] = pd.to_datetime(df_r14[col], errors='coerce').apply(
                lambda x: x if pd.notna(x) and x >= valid_from else pd.NaT)
    df_r14['Training Required? '] = df_r14['Training Required? '].fillna('Not Set')
    df_r14['CDL Name']             = df_r14['CDL Name'].fillna('Unassigned')
    df_r14['Final Overall Status'] = df_r14['Final Overall Status'].fillna('Not Set')
    def map_pg_r14(p): return 'CX' if str(p).startswith('CX') else p
    df_r14['Pillar Group'] = df_r14['Pillar'].apply(map_pg_r14)
    df_r14_nf    = df_r14[(df_r14['Feature Category'] != 'Unboxing') & (df_r14['Training Required? '] == 'Yes')].copy()
    df_r14_unbox = df_r14[(df_r14['Feature Category'] == 'Unboxing') & (df_r14['Training Required? '] == 'Yes')].copy()
    df_r14_all   = pd.concat([df_r14_nf, df_r14_unbox], ignore_index=True)
    video_filter = ['Single NF', 'Combined Primary NF']
    fc_col = 'Feature Classification'; sr_col = 'CDL Daily Status Update'
    pillar_order = ['HCM', 'ERP', 'SCM', 'PRC', 'CX']

    # ── Blocked resolution values ──
    blocked_resolution_values = [
        'Dropped ( Extensive Setup)',
        'Dropped ( Bug)',
        'Dropped ( Env does not support)',
        'Requires D01 Pod',
        'Requires D02 Pod'
    ]

    rows = []
    all_cdls = df_r14_all[df_r14_all['CDL Name'] != 'Unassigned']['CDL Name'].unique()
    for cdl in all_cdls:
        cdl_all   = df_r14_all[df_r14_all['CDL Name'] == cdl]
        cdl_nf    = df_r14_nf[df_r14_nf['CDL Name'] == cdl]
        cdl_unbox = df_r14_unbox[df_r14_unbox['CDL Name'] == cdl]
        cdl_orig  = df_r14[df_r14['CDL Name'] == cdl]

        pillar        = cdl_all['Pillar Group'].iloc[0] if len(cdl_all) > 0 else 'Unknown'
        nf_count      = len(cdl_nf)
        unbox_count   = len(cdl_unbox)
        total_count   = nf_count + unbox_count

        nf_video_count    = len(cdl_nf[cdl_nf[fc_col].isin(video_filter)]) if fc_col in cdl_nf.columns else nf_count
        nf_unclassified   = len(cdl_nf[cdl_nf[fc_col].isna()]) if fc_col in cdl_nf.columns else 0
        unbox_video_count = unbox_count
        total_video_count = nf_video_count + unbox_video_count
        nf_video_display  = f'{nf_video_count} (+{nf_unclassified} unclassified)' if nf_unclassified > 0 else str(nf_video_count)

        rec_dates  = cdl_all['Recording Date'].dropna()
        start_date = rec_dates.min() if len(rec_dates) > 0 else pd.NaT
        end_date   = rec_dates.max() if len(rec_dates) > 0 else pd.NaT
        start_disp = start_date.strftime('%d-%b-%Y') if pd.notna(start_date) else 'Not Set'
        end_disp   = end_date.strftime('%d-%b-%Y')   if pd.notna(end_date)   else 'Not Set'

        ga_tier = get_ga_tier_r14(
            end_date + pd.Timedelta(days=4) if pd.notna(end_date) else pd.NaT,
            ga_base_ts_r14
        )

        if pd.notna(start_date) and start_date <= today_r14:
            if pd.notna(end_date) and end_date < today_r14:
                ref_date = end_date
            else:
                ref_date = today_r14 - pd.Timedelta(days=1)
            days_elapsed_exp = max(int(np.busday_count(
                start_date.date(),
                (ref_date + pd.Timedelta(days=1)).date()
            )), 0)
            target_comp = min(days_elapsed_exp * 3, total_count)
        else:
            ref_date    = pd.NaT
            target_comp = 0

        # Completed primary videos (for bundled NF detection)
        completed_primary_videos = set(
            cdl_nf[
                (cdl_nf[fc_col] == 'Combined Primary NF') &
                (cdl_nf['Video Ready Date'].notna()) &
                (cdl_nf['Combined Feature Video Name'].notna())
            ]['Combined Feature Video Name']
        ) if fc_col in cdl_nf.columns else set()

        single_done  = len(cdl_nf[
            (cdl_nf[fc_col].isin(['Single NF']) | cdl_nf[fc_col].isna()) &
            (cdl_nf['Video Ready Date'].notna())
        ]) if fc_col in cdl_nf.columns else 0
        primary_done = len(cdl_nf[
            (cdl_nf[fc_col] == 'Combined Primary NF') &
            (cdl_nf['Video Ready Date'].notna())
        ]) if fc_col in cdl_nf.columns else 0
        bundled_done = len(cdl_nf[
            (cdl_nf[fc_col] == 'Combined Bundled NF') &
            (cdl_nf['Combined Feature Video Name'].isin(completed_primary_videos))
        ]) if completed_primary_videos else 0
        actual_nf    = single_done + primary_done + bundled_done
        actual_unbox = len(cdl_unbox[cdl_unbox['Video Ready Date'].notna()])
        actual_total = actual_nf + actual_unbox

        nf_released    = len(cdl_orig[(cdl_orig['Final Overall Status'] == 'Released (A)') & (cdl_orig['Feature Category'] != 'Unboxing')])
        unbox_released = len(cdl_orig[(cdl_orig['Final Overall Status'] == 'Released (A)') & (cdl_orig['Feature Category'] == 'Unboxing')])
        dropped  = len(cdl_orig[cdl_orig['Final Overall Status'] == 'Feature Dropped'])
        sr_count = len(cdl_orig[cdl_orig[sr_col].fillna('').str.contains('SR raised', case=False, na=False)]) if sr_col in cdl_orig.columns else 0

        days_worked = max(int(np.busday_count(start_date.date(), end_date.date())), 1) if pd.notna(start_date) and pd.notna(end_date) else 0

        if pd.notna(start_date) and pd.notna(ref_date):
            days_worked_actual = max(int(np.busday_count(
                start_date.date(),
                (ref_date + pd.Timedelta(days=1)).date()
            )), 1)
        else:
            days_worked_actual = 0
        features_per_day = round(actual_nf / days_worked_actual, 1) if days_worked_actual > 0 and actual_nf > 0 else 0.0

        if pd.notna(start_date):
            ref_end      = min(today_r14, end_date) if pd.notna(end_date) else today_r14
            days_elapsed = max(int(np.busday_count(start_date.date(), ref_end.date())), 0)
        else:
            days_elapsed = 0

        # ── is_feature_blocked defined FIRST so is_feature_completed can use it ──
        def is_feature_blocked(frow):
            issue_type_val   = str(frow.get('Type of Issue', '') or '').strip()
            issue_detail_val = str(frow.get('Details of the Issue faced ', '') or '').strip()
            final_status_val = str(frow.get('Final Overall Status', '') or '').strip()
            resolution_val   = str(frow.get('Issue Resolution Outcome', '') or '').strip()
            return (
                (issue_type_val   not in ['', 'nan', 'none', 'NaN']) or
                (issue_detail_val not in ['', 'nan', 'none', 'NaN']) or
                (final_status_val == 'Feature Dropped') or
                (resolution_val   in blocked_resolution_values)
            )

        # ── Combined Bundled NF with active issue is NOT treated as complete ──
        def is_feature_completed(frow):
            fc   = str(frow.get(fc_col, '') or '').strip()
            vrd  = frow.get('Video Ready Date', None)
            cfvn = str(frow.get('Combined Feature Video Name', '') or '').strip()
            cat  = str(frow.get('Feature Category', '') or '').strip()
            if cat == 'Unboxing': return pd.notna(vrd)
            if fc in ['Single NF', 'Combined Primary NF']: return pd.notna(vrd)
            if fc == 'Combined Bundled NF':
                # If blocked by an issue, do NOT treat as complete even if primary is done
                if is_feature_blocked(frow): return False
                return cfvn in completed_primary_videos
            if fc in ['', 'nan', 'none', 'NaN']: return pd.notna(vrd)
            return False

        # ── blocked split into NF and Unboxing ──
        blocked_nf_count    = 0
        blocked_unbox_count = 0

        for _, frow in cdl_nf.iterrows():
            if is_feature_completed(frow): continue
            if is_feature_blocked(frow): blocked_nf_count += 1

        for _, frow in cdl_unbox.iterrows():
            if is_feature_completed(frow): continue
            if is_feature_blocked(frow): blocked_unbox_count += 1

        blocked_count = blocked_nf_count + blocked_unbox_count

        # ── delayed split into NF and Unboxing ──
        total_delayed   = max(target_comp - actual_total - blocked_count, 0)
        pending_nf      = max(nf_count - actual_nf - blocked_nf_count, 0)
        pending_unbox   = max(unbox_count - actual_unbox - blocked_unbox_count, 0)
        if total_delayed > 0:
            delayed_nf_count    = min(total_delayed, pending_nf)
            delayed_unbox_count = max(total_delayed - delayed_nf_count, 0)
            delayed_unbox_count = min(delayed_unbox_count, pending_unbox)
        else:
            delayed_nf_count    = 0
            delayed_unbox_count = 0
        delayed_count = total_delayed

        # ── FIX (status labels):
        #    1. 'Complete' now means ALL features are actually done — not just "on pace".
        #    2. If target_comp == 0 (recording started today / after a weekend, so nothing
        #       is due yet), show 'In Progress (No videos due yet)' instead of a misleading
        #       'Complete' with '✅ Completed: 0'.
        #    3. Meeting the 3/day pace (but not fully done) is now labelled 'On Track'. ──
        if not pd.notna(start_date):                          feat_status_label = 'Recording Date Not Set'
        elif start_date > today_r14:                          feat_status_label = 'Yet To Start'
        elif total_count > 0 and actual_total >= total_count: feat_status_label = 'Complete'
        elif target_comp == 0:                                feat_status_label = 'In Progress (No videos due yet)'
        elif actual_total >= target_comp:                     feat_status_label = 'On Track'
        elif delayed_count == 0 and blocked_count > 0:        feat_status_label = 'On Track (Issues Blocking)'
        elif delayed_count == 0:                              feat_status_label = 'On Track'
        else:                                                 feat_status_label = 'Delayed'

        def fmt_completed():
            if actual_unbox > 0:
                return f'✅ Completed: {actual_nf} NF | {actual_unbox} Unboxing'
            return f'✅ Completed: {actual_nf}'

        def fmt_blocked():
            if blocked_nf_count > 0 and blocked_unbox_count > 0:
                return f'⚠️ Blocked by Issues: {blocked_nf_count} NF | {blocked_unbox_count} Unboxing'
            elif blocked_nf_count > 0:
                return f'⚠️ Blocked by Issues: {blocked_nf_count} NF'
            elif blocked_unbox_count > 0:
                return f'⚠️ Blocked by Issues: {blocked_unbox_count} Unboxing'
            return ''

        def fmt_delayed():
            if delayed_nf_count > 0 and delayed_unbox_count > 0:
                return f'🔴 Delayed: {delayed_nf_count} NF | {delayed_unbox_count} Unboxing'
            elif delayed_nf_count > 0:
                return f'🔴 Delayed: {delayed_nf_count} NF'
            elif delayed_unbox_count > 0:
                return f'🔴 Delayed: {delayed_unbox_count} Unboxing'
            return ''

        # ── FIX: handle the new labels in the display string ──
        if feat_status_label in ['Recording Date Not Set', 'Yet To Start', 'In Progress (No videos due yet)']:
            feat_status = feat_status_label
        elif feat_status_label in ['Complete', 'On Track']:
            feat_status = f'{feat_status_label}\n{fmt_completed()}'
        elif feat_status_label == 'On Track (Issues Blocking)':
            parts = [fmt_completed()]
            blocked_str = fmt_blocked()
            if blocked_str: parts.append(blocked_str)
            feat_status = 'On Track (Issues Blocking)\n' + ' | '.join(parts)
        else:
            parts = [fmt_completed()]
            blocked_str = fmt_blocked()
            delayed_str = fmt_delayed()
            if blocked_str: parts.append(blocked_str)
            if delayed_str: parts.append(delayed_str)
            feat_status = 'Delayed\n' + ' | '.join(parts)

        actual_videos = len(cdl_nf[
            (cdl_nf['Video Ready Date'].notna()) &
            (cdl_nf[fc_col].isin(video_filter))
        ]) if fc_col in cdl_nf.columns else 0
        not_set_count = len(cdl_nf[cdl_nf[fc_col].isna()]) if fc_col in cdl_nf.columns else 0
        if nf_video_count == 0 and not_set_count == 0:
            vid_status = 'Feature Classification Not Set'
        elif nf_video_count == 0 and not_set_count > 0:
            vid_status = f'⚠️ Classification Not Set: {not_set_count}'
        else:
            vid_status = f'{actual_videos} out of {nf_video_count}'
            if not_set_count > 0: vid_status += f'\n⚠️ Classification Not Set: {not_set_count}'

        issue_parts = []
        if 'Details of the Issue faced ' in cdl_orig.columns:
            for _, frow in cdl_orig.iterrows():
                issue_detail = str(frow.get('Details of the Issue faced ', '') or '').strip()
                gcc_val      = str(frow.get('Feature GCC', '') or '').strip()
                issue_type   = str(frow.get('Type of Issue', '') or '').strip()
                if issue_detail and issue_detail.lower() not in ['nan', 'none', '']:
                    prefix   = gcc_val if gcc_val and gcc_val.lower() not in ['nan','none',''] else 'Unknown'
                    type_str = f"{issue_type}: " if issue_type and issue_type.lower() not in ['nan','none',''] else ''
                    issue_parts.append(f"{prefix}: {type_str}{issue_detail}")
        issues_consolidation = '\n'.join(issue_parts) if issue_parts else ''

        consolidation_parts = []
        if sr_col in cdl_orig.columns:
            for _, frow in cdl_orig.iterrows():
                if is_feature_completed(frow): continue
                status_val   = str(frow.get(sr_col, '') or '').strip()
                feature_name = str(frow.get('Feature', '') or '').strip()
                prefix       = ' '.join(feature_name.split()[:3]) if feature_name else 'Feature'
                if status_val and status_val.lower() not in ['nan','none','']:
                    consolidation_parts.append(f"{prefix}: {status_val}")
                else:
                    issue_type   = str(frow.get('Type of Issue', '') or '').strip()
                    issue_detail = str(frow.get('Details of the Issue faced ', '') or '').strip()
                    response     = str(frow.get('Response Received', '') or '').strip()
                    resolution   = str(frow.get('Issue Resolution Outcome', '') or '').strip()
                    issue_parts_con = []
                    if issue_type   and issue_type.lower()   not in ['nan','none','']: issue_parts_con.append(issue_type)
                    if issue_detail and issue_detail.lower() not in ['nan','none','']: issue_parts_con.append(issue_detail[:200])
                    if response     and response.lower()     not in ['nan','none','']: issue_parts_con.append(f"Response: {response[:150]}")
                    if resolution   and resolution.lower()   not in ['nan','none','']: issue_parts_con.append(f"Resolved: {resolution[:150]}")
                    if issue_parts_con: consolidation_parts.append(f"{prefix}: {' | '.join(issue_parts_con)}")
        seen = set(); unique_parts = []
        for p in consolidation_parts:
            if p not in seen: seen.add(p); unique_parts.append(p)
        consolidation = '\n'.join(unique_parts) if unique_parts else ''

        rows.append({
            'Pillar': pillar, 'CDL': cdl,
            '# of NF to be developed (Including Unboxing)': total_count,
            '# NF': nf_count, '# Unboxing': unbox_count,
            '# of NF Videos to be developed (Incl. Unboxing)': total_video_count,
            '# NF Videos': nf_video_display, '# Unboxing Videos': unbox_video_count,
            'NF Dev Start Date': start_disp, 'NF Dev End Date': end_disp, 'GA+X': ga_tier,
            'Expected Completion': target_comp, 'Actual NF Completion': actual_nf,
            'Actual Unboxing Completion': actual_unbox,
            'Feature Count Status (3/Day rule)': feat_status,
            'No Of Videos Recorded': vid_status,
            '# NF Released to Production': nf_released,
            '# Unboxing Released to Production': unbox_released,
            '# Features Dropped': dropped, '# SR Raised': sr_count,
            'Details of Issue Faced': issues_consolidation,
            'No of Days Worked': days_worked,
            'Number of Features Developed / Day': features_per_day,
            'CDL Status Consolidation on Pending Features': consolidation,
        })

    df_r14_out = pd.DataFrame(rows)
    pillar_order_map = {p: i for i, p in enumerate(pillar_order)}
    df_r14_out['_sort'] = df_r14_out['Pillar'].map(lambda x: pillar_order_map.get(x, 99))
    df_r14_out = df_r14_out.sort_values(['_sort','CDL']).drop(columns=['_sort']).reset_index(drop=True)

    numeric_cols_r14 = [
        '# of NF to be developed (Including Unboxing)', '# NF', '# Unboxing',
        '# of NF Videos to be developed (Incl. Unboxing)', '# Unboxing Videos',
        'Expected Completion', 'Actual NF Completion', 'Actual Unboxing Completion',
        '# NF Released to Production', '# Unboxing Released to Production',
        '# Features Dropped', '# SR Raised', 'No of Days Worked'
    ]
    total_r14 = {'Pillar': 'Total', 'CDL': ''}
    for c in df_r14_out.columns:
        if c in numeric_cols_r14: total_r14[c] = df_r14_out[c].sum()
        elif c not in ['Pillar', 'CDL']: total_r14[c] = ''
    df_r14_out = pd.concat([df_r14_out, pd.DataFrame([total_r14])], ignore_index=True)

    st.markdown('<div class="section-header">📊 CDL Status Preview</div>', unsafe_allow_html=True)

    def style_r14(row):
        styles = [''] * len(row); cols = list(row.index)
        if row['Pillar'] == 'Total': return ['font-weight: bold; background-color: #f0f0f0'] * len(row)
        if 'Feature Count Status (3/Day rule)' in cols:
            val  = row['Feature Count Status (3/Day rule)']
            idx  = cols.index('Feature Count Status (3/Day rule)')
            first_line = str(val).split('\n')[0].strip() if val else ''
            if first_line == 'Complete':
                styles[idx] = 'color: #10b981; font-weight: bold; text-align: left; white-space: pre-wrap'
            elif first_line == 'On Track':
                # ── FIX: new 'On Track' label (meeting 3/day pace) ──
                styles[idx] = 'color: #0ea5e9; font-weight: bold; text-align: left; white-space: pre-wrap'
            elif first_line == 'Delayed':
                styles[idx] = 'color: red; font-weight: bold; text-align: left; white-space: pre-wrap'
            elif first_line == 'On Track (Issues Blocking)':
                styles[idx] = 'color: #f59e0b; font-weight: bold; text-align: left; white-space: pre-wrap'
            elif first_line == 'In Progress (No videos due yet)':
                # ── FIX: new 'In Progress' label (recording just started, nothing due yet) ──
                styles[idx] = 'color: #7c3aed; font-weight: bold; text-align: left; white-space: pre-wrap'
            elif first_line == 'Yet To Start':
                styles[idx] = 'color: #4f46e5; font-weight: bold; text-align: left; white-space: pre-wrap'
            elif first_line == 'Recording Date Not Set':
                styles[idx] = 'color: #94a3b8; font-weight: bold; text-align: left; white-space: pre-wrap'
        if 'No Of Videos Recorded' in cols:
            idx = cols.index('No Of Videos Recorded'); val = row['No Of Videos Recorded']
            if val == 'Feature Classification Not Set':
                styles[idx] = 'color: #94a3b8; font-style: italic; text-align: left'
            else:
                styles[idx] = 'color: #1a1a2e; font-weight: bold; text-align: left'
        return styles

    def col_config_r14(df):
        numeric_cols = [
            '# of NF to be developed (Including Unboxing)', '# NF', '# Unboxing',
            '# of NF Videos to be developed (Incl. Unboxing)', '# NF Videos',
            '# Unboxing Videos', 'Actual NF Completion', 'Actual Unboxing Completion',
            '# NF Released to Production', '# Unboxing Released to Production',
            '# Features Dropped', '# SR Raised', 'No of Days Worked',
            'Number of Features Developed / Day'
        ]
        config = {}
        for col in df.columns:
            if col in numeric_cols: config[col] = st.column_config.NumberColumn(col, width='small')
            else: config[col] = st.column_config.TextColumn(col, width='small')
        return config

    styled_r14 = (
        df_r14_out.style
        .apply(style_r14, axis=1)
        .set_properties(**{'text-align': 'left', 'white-space': 'pre-wrap'})
        .set_table_styles([
            {'selector': 'th', 'props': [('text-align', 'left'), ('white-space', 'pre-wrap')]},
            {'selector': 'td', 'props': [('text-align', 'left'), ('white-space', 'pre-wrap')]},
        ])
        .hide(axis='index')
    )
    st.dataframe(styled_r14, use_container_width=True, column_config=col_config_r14(df_r14_out))

    st.markdown("""<div style="background-color:#f8f9fa; padding:10px 14px; border-radius:6px;
        border-left:4px solid #4f46e5; margin-top:8px; font-size:0.82rem; color:#444;">
        <b>📌 Feature Count Status Guide:</b><br>
        <span style="color:#10b981; font-weight:bold;">● Complete</span> — ALL assigned features are done &nbsp;|&nbsp;
        <span style="color:#0ea5e9; font-weight:bold;">● On Track</span> — Meeting the 3/day pace (work still remaining) &nbsp;|&nbsp;
        <span style="color:#f59e0b; font-weight:bold;">● On Track (Issues Blocking)</span> — All unblocked features done; remaining blocked by issues/dropped &nbsp;|&nbsp;
        <span style="color:red; font-weight:bold;">● Delayed</span> — Features pending with no logged issue &nbsp;|&nbsp;
        <span style="color:#7c3aed; font-weight:bold;">● In Progress (No videos due yet)</span> — Recording just started; first videos not yet due per 3/day rule &nbsp;|&nbsp;
        <span style="color:#4f46e5; font-weight:bold;">● Yet To Start</span> — Recording start date is in the future &nbsp;|&nbsp;
        <span style="color:#94a3b8; font-weight:bold;">● Recording Date Not Set</span> — No recording date entered
    </div>""", unsafe_allow_html=True)

    def generate_r14_excel(df):
        wb = Workbook(); ws = wb.active; ws.title = "CDL Status Report"
        thin   = Side(style='thin', color='AAAAAA')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left   = Alignment(horizontal='left',   vertical='top',    wrap_text=True)
        ws.merge_cells(f'A1:{get_column_letter(len(df.columns))}1')
        ws['A1'] = f'CDL Status Report — {today_r14.strftime("%d-%b-%Y")}'
        ws['A1'].font      = Font(name='Arial', bold=True, size=13, color='FFFFFF')
        ws['A1'].fill      = PatternFill('solid', start_color='1F3864')
        ws['A1'].alignment = center
        ws.row_dimensions[1].height = 25
        hdr_fill = PatternFill('solid', start_color='2E75B6')
        hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        for c_idx, col in enumerate(df.columns, 1):
            cell = ws.cell(row=2, column=c_idx, value=col)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = center; cell.border = border
        ws.row_dimensions[2].height = 40
        pillar_fills = {
            'HCM': PatternFill('solid', start_color='E2EFDA'),
            'ERP': PatternFill('solid', start_color='DDEBF7'),
            'SCM': PatternFill('solid', start_color='FCE4D6'),
            'PRC': PatternFill('solid', start_color='EDE7F6'),
            'CX':  PatternFill('solid', start_color='FFF9C4')
        }
        complete_fill     = PatternFill('solid', start_color='C6EFCE')
        on_track_pace_fill = PatternFill('solid', start_color='D9F2FB')   # ── FIX: fill for 'On Track' ──
        delayed_fill      = PatternFill('solid', start_color='FFC7CE')
        on_track_fill     = PatternFill('solid', start_color='FFEB9C')
        in_progress_fill  = PatternFill('solid', start_color='EDE7F6')    # ── FIX: fill for 'In Progress' ──
        yet_to_start_fill = PatternFill('solid', start_color='DDEBF7')
        no_date_fill      = PatternFill('solid', start_color='F2F2F2')
        total_fill        = PatternFill('solid', start_color='D9D9D9')
        for r_idx, row in enumerate(df.itertuples(index=False), 3):
            pillar_val = str(row.Pillar)
            is_total   = pillar_val == 'Total'
            row_fill   = total_fill if is_total else pillar_fills.get(pillar_val, PatternFill('solid', start_color='FFFFFF'))
            for c_idx, val in enumerate(row, 1):
                cell     = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = border
                col_name = df.columns[c_idx - 1]
                if is_total:
                    cell.fill      = total_fill
                    cell.font      = Font(name='Arial', bold=True, size=9)
                    cell.alignment = center
                elif col_name == 'Feature Count Status (3/Day rule)':
                    first_line = str(val).split('\n')[0].strip() if val else ''
                    if first_line == 'Complete':
                        cell.fill = complete_fill
                        cell.font = Font(name='Arial', size=9, bold=True, color='375623')
                    elif first_line == 'On Track':
                        # ── FIX: new 'On Track' label ──
                        cell.fill = on_track_pace_fill
                        cell.font = Font(name='Arial', size=9, bold=True, color='1F4E79')
                    elif first_line == 'Delayed':
                        cell.fill = delayed_fill
                        cell.font = Font(name='Arial', size=9, bold=True, color='9C0006')
                    elif first_line == 'On Track (Issues Blocking)':
                        cell.fill = on_track_fill
                        cell.font = Font(name='Arial', size=9, bold=True, color='7D4E00')
                    elif first_line == 'In Progress (No videos due yet)':
                        # ── FIX: new 'In Progress' label ──
                        cell.fill = in_progress_fill
                        cell.font = Font(name='Arial', size=9, bold=True, color='5B2C87')
                    elif first_line == 'Yet To Start':
                        cell.fill = yet_to_start_fill
                        cell.font = Font(name='Arial', size=9, bold=True, color='1F3864')
                    elif first_line == 'Recording Date Not Set':
                        cell.fill = no_date_fill
                        cell.font = Font(name='Arial', size=9, bold=True, color='808080')
                    else:
                        cell.fill = row_fill
                        cell.font = Font(name='Arial', size=9)
                    cell.alignment = left
                elif col_name == 'No Of Videos Recorded':
                    cell.fill      = row_fill
                    cell.font      = Font(name='Arial', size=9, bold=True,
                                         color='808080' if val == 'Feature Classification Not Set' else '1F3864')
                    cell.alignment = center
                elif col_name == 'Number of Features Developed / Day':
                    cell.fill      = row_fill
                    cell.font      = Font(name='Arial', size=9, bold=True, color='1F3864')
                    cell.alignment = center
                elif col_name in ['Details of Issue Faced', 'CDL Status Consolidation on Pending Features']:
                    cell.fill      = row_fill
                    cell.font      = Font(name='Arial', size=9)
                    cell.alignment = left
                else:
                    cell.fill      = row_fill
                    cell.font      = Font(name='Arial', size=9)
                    cell.alignment = center
            ws.row_dimensions[r_idx].height = 70 if not is_total else 25

        col_widths = {
            'Pillar': 10, 'CDL': 20,
            '# of NF to be developed (Including Unboxing)': 16,
            '# NF': 10, '# Unboxing': 10,
            '# of NF Videos to be developed (Incl. Unboxing)': 18,
            '# NF Videos': 12, '# Unboxing Videos': 14,
            'NF Dev Start Date': 14, 'NF Dev End Date': 14, 'GA+X': 10,
            'Expected Completion': 12, 'Actual NF Completion': 14,
            'Actual Unboxing Completion': 16,
            'Feature Count Status (3/Day rule)': 30,
            'No Of Videos Recorded': 20,
            '# NF Released to Production': 14,
            '# Unboxing Released to Production': 16,
            '# Features Dropped': 12, '# SR Raised': 10,
            'Details of Issue Faced': 45,
            'No of Days Worked': 12,
            'Number of Features Developed / Day': 16,
            'CDL Status Consolidation on Pending Features': 60
        }
        for c_idx, col in enumerate(df.columns, 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = col_widths.get(col, 14)
        ws.freeze_panes = 'A3'
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    excel_bytes = generate_r14_excel(df_r14_out)
    st.markdown("---")
    st.markdown('<div class="section-header">⬇️ Download Daily CDL Status Report</div>', unsafe_allow_html=True)
    st.download_button(
        f"⬇️ Download CDL Status Report — {today_r14.strftime('%d-%b-%Y')}",
        excel_bytes,
        f"CDL_Status_Report_{today_r14.strftime('%d%b%Y')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
